#!/usr/bin/env python3
"""
TG Partner Tracker — 拉取合作方群历史消息 + 进度跟踪 + 飞书同步

Usage:
  python3 tg_history.py login                    # 首次登录（输手机号+验证码）
  python3 tg_history.py groups                   # 列出你加入的所有群
  python3 tg_history.py fetch                    # 拉取配置中指定群的全部历史消息
  python3 tg_history.py fetch --since 2025-01-01 # 只拉指定日期之后的
  python3 tg_history.py export                   # 生成 Excel（按群分 sheet）
  python3 tg_history.py sync                     # 推送飞书
"""
from __future__ import annotations

import argparse
import asyncio
import json
import subprocess
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from telethon import TelegramClient
    from telethon.tl.types import Channel, Chat, User, Message
except ImportError:
    print("需要 telethon:  pip install telethon")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl:  pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
DATA_DIR = SCRIPT_DIR / "data"
SESSION_FILE = SCRIPT_DIR / "tg_session"
EXCEL_PATH = SCRIPT_DIR / "partner_history.xlsx"

FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_API = "https://open.feishu.cn/open-apis/bitable/v1"

_fs_token: str | None = None
_fs_token_exp: float = 0.0


# ═══════════════════════════════════════════════════════════════════
#  Config
# ═══════════════════════════════════════════════════════════════════

def load_config() -> dict:
    if not CONFIG_PATH.is_file():
        print(f"❌ 找不到 {CONFIG_PATH}"); sys.exit(1)
    return json.loads(CONFIG_PATH.read_text("utf-8"))


def get_client(cfg: dict) -> TelegramClient:
    api_id = int(cfg["api_id"])
    api_hash = cfg["api_hash"]
    return TelegramClient(str(SESSION_FILE), api_id, api_hash)


# ═══════════════════════════════════════════════════════════════════
#  Login
# ═══════════════════════════════════════════════════════════════════

async def do_login(cfg: dict):
    client = get_client(cfg)
    await client.start()
    me = await client.get_me()
    print(f"✅ 登录成功: {me.first_name} (@{me.username}) ID={me.id}")
    await client.disconnect()


# ═══════════════════════════════════════════════════════════════════
#  List groups
# ═══════════════════════════════════════════════════════════════════

async def do_groups(cfg: dict):
    client = get_client(cfg)
    await client.start()

    print("\n📋 你加入的群组/频道:\n")
    print(f"  {'ID':<16} {'类型':<10} {'成员':<8} 名称")
    print("  " + "─" * 60)

    async for dialog in client.iter_dialogs():
        entity = dialog.entity
        if isinstance(entity, (Channel, Chat)):
            members = getattr(entity, "participants_count", None) or "?"
            is_channel = isinstance(entity, Channel)
            kind = "频道" if (is_channel and entity.broadcast) else "超级群" if is_channel else "群"
            print(f"  {dialog.id:<16} {kind:<10} {str(members):<8} {dialog.name}")

    await client.disconnect()
    print("\n把要跟踪的群 ID 填入 config.json 的 partner_groups 里。")


# ═══════════════════════════════════════════════════════════════════
#  Fetch history
# ═══════════════════════════════════════════════════════════════════

async def do_fetch(cfg: dict, since: str | None = None):
    client = get_client(cfg)
    await client.start()

    groups = cfg.get("partner_groups") or []
    if not groups:
        print("❌ config.json 里 partner_groups 为空")
        print("   先运行 python3 tg_history.py groups 查看群列表")
        await client.disconnect()
        return

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    since_dt = None
    if since:
        since_dt = datetime.strptime(since, "%Y-%m-%d").replace(
            tzinfo=timezone.utc)

    for group_cfg in groups:
        gid = group_cfg["id"]
        name = group_cfg.get("name", str(gid))
        print(f"\n📥 拉取 [{name}] (ID: {gid}) ...")

        try:
            entity = await client.get_entity(int(gid))
        except Exception as e:
            print(f"  ❌ 无法访问: {e}")
            continue

        out_file = DATA_DIR / f"group_{gid}.jsonl"
        existing_ids: set[str] = set()
        if out_file.is_file():
            with out_file.open("r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line:
                        try:
                            existing_ids.add(str(json.loads(line).get("msg_id", "")))
                        except json.JSONDecodeError:
                            pass

        count = 0
        skipped = 0
        with out_file.open("a", encoding="utf-8") as f:
            async for msg in client.iter_messages(entity, reverse=True,
                                                   offset_date=since_dt):
                if not isinstance(msg, Message) or not msg.text:
                    continue

                mid = str(msg.id)
                if mid in existing_ids:
                    skipped += 1
                    continue

                sender_name = ""
                sender_un = ""
                sender_id = ""
                if msg.sender:
                    if isinstance(msg.sender, User):
                        sender_name = f"{msg.sender.first_name or ''} {msg.sender.last_name or ''}".strip()
                        sender_un = msg.sender.username or ""
                        sender_id = str(msg.sender.id)
                    else:
                        sender_name = getattr(msg.sender, "title", "") or ""
                        sender_id = str(msg.sender.id)

                rec = {
                    "msg_id": mid,
                    "chat_id": str(gid),
                    "chat_title": name,
                    "sender_id": sender_id,
                    "sender_username": sender_un,
                    "sender_name": sender_name,
                    "text": msg.text,
                    "timestamp": msg.date.strftime("%Y-%m-%d %H:%M:%S UTC"),
                    "reply_to": str(msg.reply_to_msg_id) if msg.reply_to_msg_id else "",
                }
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                count += 1

                if count % 500 == 0:
                    print(f"  ... {count} messages")

        print(f"  ✅ {count} new, {skipped} skipped (already fetched)")

    await client.disconnect()
    print(f"\n✅ 全部完成。数据保存在 {DATA_DIR}/")


# ═══════════════════════════════════════════════════════════════════
#  Export → Excel (one sheet per group)
# ═══════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")

COLS = [
    ("时间",      20),
    ("发送者",    16),
    ("用户名",    14),
    ("消息内容",  60),
    ("msg_id",    12),
    ("进度备注",  30),
]


def do_export(cfg: dict):
    groups = cfg.get("partner_groups") or []
    if not groups:
        print("❌ partner_groups 为空"); return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_msgs = 0
    for group_cfg in groups:
        gid = group_cfg["id"]
        name = group_cfg.get("name", str(gid))
        safe_name = name[:31].replace("/", "-").replace("\\", "-")

        jsonl_path = DATA_DIR / f"group_{gid}.jsonl"
        if not jsonl_path.is_file():
            print(f"  ⚠️ [{name}] 无数据，跳过")
            continue

        msgs = []
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        msgs.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass

        ws = wb.create_sheet(title=safe_name)

        for ci, (label, width) in enumerate(COLS, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
        ws.freeze_panes = "A2"

        for ri, m in enumerate(msgs, 2):
            row = [
                m.get("timestamp", ""),
                m.get("sender_name", ""),
                f"@{m['sender_username']}" if m.get("sender_username") else "",
                m.get("text", ""),
                m.get("msg_id", ""),
                "",
            ]
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = CELL_FONT
                cell.alignment = WRAP

        total_msgs += len(msgs)
        print(f"  📄 [{name}] {len(msgs)} messages")

    # Summary sheet
    ws_sum = wb.create_sheet("汇总", 0)
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 20
    ws_sum.column_dimensions["D"].width = 20
    ws_sum.column_dimensions["E"].width = 40

    headers = ["合作方群", "消息数", "最早消息", "最新消息", "进度备注"]
    for ci, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for ri, group_cfg in enumerate(groups, 2):
        gid = group_cfg["id"]
        name = group_cfg.get("name", str(gid))
        jsonl_path = DATA_DIR / f"group_{gid}.jsonl"

        msg_count = 0
        earliest = ""
        latest = ""
        if jsonl_path.is_file():
            with jsonl_path.open("r", encoding="utf-8") as f:
                lines = [l.strip() for l in f if l.strip()]
            msg_count = len(lines)
            if lines:
                try:
                    first = json.loads(lines[0])
                    last = json.loads(lines[-1])
                    earliest = first.get("timestamp", "")
                    latest = last.get("timestamp", "")
                except json.JSONDecodeError:
                    pass

        ws_sum.cell(row=ri, column=1, value=name).font = CELL_FONT
        ws_sum.cell(row=ri, column=2, value=msg_count).font = CELL_FONT
        ws_sum.cell(row=ri, column=3, value=earliest).font = CELL_FONT
        ws_sum.cell(row=ri, column=4, value=latest).font = CELL_FONT
        ws_sum.cell(row=ri, column=5, value="").font = CELL_FONT

    wb.save(EXCEL_PATH)
    print(f"\n✅ 导出 {total_msgs} 条消息 → {EXCEL_PATH}")
    print(f"   汇总页可以填写「进度备注」，每个群一个 sheet")
    subprocess.run(["open", str(EXCEL_PATH)], check=False)


# ═══════════════════════════════════════════════════════════════════
#  Feishu sync
# ═══════════════════════════════════════════════════════════════════

def _curl(method: str, url: str, headers: dict,
          body: dict | None = None) -> dict:
    cmd = ["curl", "-s", "-X", method, url]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    if body is not None:
        cmd += ["-d", json.dumps(body)]
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    if not p.stdout.strip():
        return {}
    return json.loads(p.stdout)


def fs_token(cfg: dict) -> str:
    global _fs_token, _fs_token_exp
    now = time.time()
    if _fs_token and now < _fs_token_exp - 120:
        return _fs_token
    fs = cfg["feishu"]
    r = _curl("POST", FEISHU_TOKEN_URL,
              {"Content-Type": "application/json"},
              {"app_id": fs["app_id"], "app_secret": fs["app_secret"]})
    if r.get("code") != 0:
        raise RuntimeError(f"飞书 token: {r}")
    _fs_token = r["tenant_access_token"]
    _fs_token_exp = now + int(r.get("expire", 7200))
    return _fs_token


def fs_create(token: str, app_token: str, table_id: str, fields: dict):
    url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records"
    r = _curl("POST", url,
              {"Authorization": f"Bearer {token}",
               "Content-Type": "application/json"},
              {"fields": fields})
    if r.get("code") != 0:
        raise RuntimeError(f"飞书写入: {r}")


def _read_group_msgs(gid: str) -> list[dict]:
    jsonl_path = DATA_DIR / f"group_{gid}.jsonl"
    if not jsonl_path.is_file():
        return []
    msgs = []
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    msgs.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return msgs


MY_USERNAMES = {"jchen_web3", "henghengcat_bot", "jchen", "hazel"}


def _build_partner_summary(name: str, gid: str, msgs: list[dict]) -> dict:
    if not msgs:
        return {
            "项目名称": name,
            "总消息数": 0, "我的回复数": 0,
            "最早消息": "", "最新消息": "",
            "最近活动摘要": "无消息",
            "最近一周进度": "",
            "我的每日回复": "",
            "上下文": "",
            "群ID": gid,
        }

    total = len(msgs)
    earliest = msgs[0].get("timestamp", "") if msgs else ""
    latest = msgs[-1].get("timestamp", "") if msgs else ""

    my_msgs = [m for m in msgs
               if (m.get("sender_username") or "").lower() in MY_USERNAMES]
    my_count = len(my_msgs)

    now = datetime.now(timezone.utc)
    week_ago = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    recent_msgs = [m for m in msgs if m.get("timestamp", "") >= week_ago]

    recent_summary_parts = []
    for m in recent_msgs[-20:]:
        sender = m.get("sender_name", "?")
        text = (m.get("text") or "")[:100]
        recent_summary_parts.append(f"[{sender}]: {text}")
    recent_summary = "\n".join(recent_summary_parts[-10:]) if recent_summary_parts else "最近一周无消息"

    week_progress = f"本周 {len(recent_msgs)} 条消息"
    if recent_msgs:
        week_my = [m for m in recent_msgs
                   if (m.get("sender_username") or "").lower() in MY_USERNAMES]
        week_progress += f"，其中我的 {len(week_my)} 条"

    my_daily: dict[str, int] = {}
    for m in my_msgs:
        day = m.get("timestamp", "")[:10]
        if day:
            my_daily[day] = my_daily.get(day, 0) + 1
    recent_days = sorted(my_daily.keys())[-7:]
    daily_str = " | ".join(f"{d}: {my_daily[d]}" for d in recent_days) if recent_days else "无"

    last_5 = msgs[-5:]
    context = "\n".join(
        f"[{m.get('sender_name','?')}] {(m.get('text') or '')[:150]}"
        for m in last_5)

    return {
        "项目名称": name,
        "总消息数": total,
        "我的回复数": my_count,
        "最早消息": earliest,
        "最新消息": latest,
        "最近活动摘要": recent_summary[:2000],
        "最近一周进度": week_progress,
        "我的每日回复": daily_str[:2000],
        "上下文": context[:2000],
        "群ID": gid,
        "recent_week_count": len(recent_msgs),
    }


def _llm_translate(text: str, api_key: str) -> str:
    if not text or not api_key or len(text.strip()) < 5:
        return text
    body = {
        "model": "openai/gpt-4o-mini",
        "messages": [
            {"role": "system", "content":
             "Translate the following chat messages to Chinese. "
             "Format: keep each original line, add a Chinese translation below it. "
             "Like:\n[Sender]: original message\n→ 中文翻译\n\n"
             "If a line is already in Chinese, keep it as-is without adding translation. "
             "Be concise. Do NOT add explanations."},
            {"role": "user", "content": text},
        ],
        "temperature": 0.3,
    }
    r = _curl("POST", "https://openrouter.ai/api/v1/chat/completions",
              {"Authorization": f"Bearer {api_key}",
               "Content-Type": "application/json"}, body)
    try:
        return r["choices"][0]["message"]["content"].strip()
    except (KeyError, IndexError):
        return text


def _llm_summarize(name: str, raw_msgs: str, week_progress: str,
                   my_count: int, total: int, api_key: str) -> str:
    if not api_key or not raw_msgs or raw_msgs == "最近一周无消息":
        return raw_msgs
    body = {
        "model": "openai/gpt-4o-mini",
        "messages": [
            {"role": "system", "content":
             "你是 DevRel，负责跟踪合作方群的进展。"
             "根据提供的最近聊天记录，用 2-4 句简洁中文总结当前合作状态。"
             "包含：在推进什么事、谁在活跃、有什么卡点或待办。"
             "不要列举每条消息，要提炼要点。"},
            {"role": "user", "content":
             f"合作方: {name}\n"
             f"本周统计: {week_progress}，我的回复 {my_count} 条，历史总消息 {total} 条\n\n"
             f"最近聊天记录:\n{raw_msgs}"},
        ],
        "temperature": 0.4,
    }
    r = _curl("POST", "https://openrouter.ai/api/v1/chat/completions",
              {"Authorization": f"Bearer {api_key}",
               "Content-Type": "application/json"}, body)
    try:
        return r["choices"][0]["message"]["content"].strip()
    except (KeyError, IndexError):
        return raw_msgs


def do_sync(cfg: dict):
    fs = cfg.get("feishu") or {}
    if not fs.get("enabled"):
        print("⚠️ feishu.enabled=false"); return

    kanban_tid = fs.get("kanban_table_id")
    if not kanban_tid:
        print("⚠️ kanban_table_id 未配置"); return

    tk = fs_token(cfg)
    app_token = fs["app_token"]
    groups = cfg.get("partner_groups") or []

    api_key = cfg.get("openrouter_api_key", "").strip()

    print(f"📤 同步 {len(groups)} 个合作方到飞书看板 ...\n")
    ok = 0
    for idx, group_cfg in enumerate(groups):
        gid = group_cfg["id"]
        name = group_cfg.get("name", str(gid))
        msgs = _read_group_msgs(gid)
        summary = _build_partner_summary(name, gid, msgs)

        week_count = summary.get("recent_week_count", 0)
        if week_count >= 10:
            priority = "P0-紧急"
            status = "正在进行"
        elif week_count >= 1:
            priority = "P1-高"
            status = "正在进行"
        elif summary["总消息数"] > 0:
            priority = "P2-中"
            status = "沉寂中"
        else:
            priority = "P3-低"
            status = "沉寂中"

        recent_summary = summary["最近活动摘要"]
        context = summary["上下文"]
        if api_key and recent_summary and recent_summary != "最近一周无消息":
            print(f"  🔤 [{idx+1}/{len(groups)}] AI 总结中 ...")
            recent_summary = _llm_summarize(
                name, recent_summary, summary["最近一周进度"],
                summary["我的回复数"], summary["总消息数"], api_key)
            context = _llm_translate(context, api_key)

        fields = {
            "项目名称":       summary["项目名称"],
            "优先级":         priority,
            "进度状态":       status,
            "最近活动摘要":   recent_summary[:2000],
            "我的回复数":     summary["我的回复数"],
            "上下文":         context[:2000],
            "最近一周进度":   summary["最近一周进度"],
            "我的每日回复":   summary["我的每日回复"],
            "总消息数":       summary["总消息数"],
            "最早消息":       summary["最早消息"],
            "最新消息":       summary["最新消息"],
            "群ID":           summary["群ID"],
            "Email数据":      "",
            "Google Meeting数据": "",
            "当前进度":       "",
        }

        try:
            fs_create(tk, app_token, kanban_tid, fields)
            ok += 1
            print(f"  ✅ [{name}] {summary['总消息数']} msgs, 我的 {summary['我的回复数']}")
        except Exception as e:
            print(f"  ❌ [{name}] {e}")

    print(f"\n✅ 完成: {ok}/{len(groups)} 合作方已推送飞书看板")


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description="TG Partner Tracker")
    sub = p.add_subparsers(dest="cmd")

    sub.add_parser("login", help="首次登录 Telegram")
    sub.add_parser("groups", help="列出你加入的所有群")

    fe = sub.add_parser("fetch", help="拉取合作方群历史消息")
    fe.add_argument("--since", help="只拉此日期之后 (YYYY-MM-DD)")

    sub.add_parser("export", help="生成 Excel")
    sub.add_parser("sync", help="推送飞书")

    args = p.parse_args()
    if not args.cmd:
        p.print_help(); return

    cfg = load_config()

    if args.cmd == "login":
        asyncio.run(do_login(cfg))
    elif args.cmd == "groups":
        asyncio.run(do_groups(cfg))
    elif args.cmd == "fetch":
        asyncio.run(do_fetch(cfg, since=args.since))
    elif args.cmd == "export":
        do_export(cfg)
    elif args.cmd == "sync":
        do_sync(cfg)


if __name__ == "__main__":
    main()
