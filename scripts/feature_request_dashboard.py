#!/usr/bin/env python3
"""
Feature Request 看板生成器

1. 读取 forum_posts.json + scored_requests.json
2. LLM 批量翻译标题 + 摘要为中文
3. 导出 Excel: 「最近7天」sheet + 「全部」sheet + 「统计」sheet
4. 可选推送飞书多维表格

用法:
  python3 feature_request_dashboard.py translate    # LLM翻译 → 缓存
  python3 feature_request_dashboard.py export       # 导出 Excel
  python3 feature_request_dashboard.py sync --table-id tblXXX   # 推飞书
  python3 feature_request_dashboard.py run          # translate + export 一键
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

SCRIPT_DIR = Path(__file__).resolve().parent
FORUM_PATH = SCRIPT_DIR / "forum_export" / "forum_posts.json"
SCORED_PATH = SCRIPT_DIR / "forum_export" / "scored_requests.json"
TRANSLATED_PATH = SCRIPT_DIR / "forum_export" / "translated_dashboard.json"
EXCEL_PATH = SCRIPT_DIR / "feature_request_dashboard.xlsx"

CONFIG_CANDIDATES = [
    SCRIPT_DIR.parent / "astrbot_plugin_dc_assistant" / "config.json",
    SCRIPT_DIR.parent / "astrbot_plugin_dc_user_collector" / "feishu_config.json",
]

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_API = "https://open.feishu.cn/open-apis/bitable/v1"

_fs_token: str | None = None
_fs_token_exp: float = 0.0

TRANSLATE_PROMPT = """\
Translate the following Discord feature request into Chinese. Output JSON only (no markdown):

{{
  "title_zh": "<Chinese title, concise>",
  "content_zh": "<Chinese summary of the request in 1-3 sentences, max 150 chars>",
  "replies_zh": "<Chinese summary of community replies in 1-2 sentences, or '无回复'>",
  "component": "<affected area: 交易/钱包/工作流/自动驾驶/聊天/UI/社区/其他>"
}}

Title: {title}
Author: {author}
Tags: {tags}
Content:
{content}

Replies ({reply_count}):
{replies}
"""


# ═══════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════

def load_config() -> dict:
    for p in CONFIG_CANDIDATES:
        if p.is_file():
            cfg = json.loads(p.read_text("utf-8"))
            return cfg
    return {}


def _curl(method: str, url: str, headers: dict,
          body: dict | None = None, timeout: int = 60) -> dict | list:
    cmd = ["curl", "-s", "-X", method, url]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    if body is not None:
        cmd += ["-d", json.dumps(body)]
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if not p.stdout.strip():
        return {}
    try:
        return json.loads(p.stdout)
    except json.JSONDecodeError:
        return {}


def load_data() -> list[dict]:
    """Merge forum_posts + scored_requests into unified records."""
    with open(FORUM_PATH, encoding="utf-8") as f:
        posts = json.load(f)

    scores_map = {}
    if SCORED_PATH.is_file():
        with open(SCORED_PATH, encoding="utf-8") as f:
            scored = json.load(f)
        for s in scored:
            scores_map[s.get("title", "")] = s

    records = []
    for post in posts:
        title = post.get("title", "")
        score = scores_map.get(title, {})

        replies_preview = ""
        for r in (post.get("replies") or [])[:5]:
            replies_preview += f"[{r.get('author','?')}]: {r.get('content','')[:150]}\n"

        records.append({
            "title": title,
            "author": post.get("author", ""),
            "tags": post.get("tags", []),
            "content": post.get("content", ""),
            "reply_count": post.get("reply_count", 0),
            "replies_preview": replies_preview,
            "created_at": post.get("created_at", ""),
            "is_archived": post.get("is_archived", False),
            "overall_score": score.get("overall_score", 0),
            "user_value": score.get("user_value", 0),
            "business_impact": score.get("business_impact", 0),
            "feasibility": score.get("feasibility", 0),
            "verdict": score.get("verdict", ""),
            "reason_zh": score.get("reason_zh", ""),
            "reason_en": score.get("reason_en", ""),
            # to be filled by translate
            "title_zh": "",
            "content_zh": "",
            "replies_zh": "",
            "component": "",
        })

    return records


# ═══════════════════════════════════════════════════════════════════
#  TRANSLATE
# ═══════════════════════════════════════════════════════════════════

def do_translate(api_key: str, model: str):
    records = load_data()

    existing = {}
    if TRANSLATED_PATH.is_file():
        existing_list = json.load(open(TRANSLATED_PATH, encoding="utf-8"))
        for r in existing_list:
            existing[r.get("title", "")] = r

    untranslated = [r for r in records if r["title"] not in existing
                    or not existing[r["title"]].get("title_zh")]

    print(f"📊 总计 {len(records)} 条, 已翻译 {len(existing)}, 待翻译 {len(untranslated)}")

    for i, rec in enumerate(untranslated):
        print(f"  [{i+1}/{len(untranslated)}] {rec['title'][:45]}...", end=" ", flush=True)

        prompt = TRANSLATE_PROMPT.format(
            title=rec["title"],
            author=rec["author"],
            tags=", ".join(rec["tags"]) or "none",
            content=(rec["content"] or "(empty)")[:1500],
            reply_count=rec["reply_count"],
            replies=rec["replies_preview"][:1000] or "(no replies)",
        )

        body = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.3,
        }
        try:
            r = _curl("POST", OPENROUTER_URL,
                       {"Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json"}, body)
            text = (r.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()
            if text.startswith("```"):
                text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            parsed = json.loads(text)
            rec["title_zh"] = parsed.get("title_zh", "")
            rec["content_zh"] = parsed.get("content_zh", "")
            rec["replies_zh"] = parsed.get("replies_zh", "")
            rec["component"] = parsed.get("component", "")
            existing[rec["title"]] = rec
            print(f"→ {rec['title_zh'][:30]}")
        except Exception as e:
            print(f"→ ❌ {e}")
            rec["title_zh"] = rec["title"]
            rec["content_zh"] = ""
            existing[rec["title"]] = rec
        time.sleep(0.8)

    # Merge back
    for rec in records:
        if rec["title"] in existing:
            cached = existing[rec["title"]]
            rec["title_zh"] = cached.get("title_zh", "")
            rec["content_zh"] = cached.get("content_zh", "")
            rec["replies_zh"] = cached.get("replies_zh", "")
            rec["component"] = cached.get("component", rec.get("component", ""))

    with open(TRANSLATED_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 翻译完成 → {TRANSLATED_PATH}")
    return records


# ═══════════════════════════════════════════════════════════════════
#  EXPORT — Excel
# ═══════════════════════════════════════════════════════════════════

VERDICT_ZH = {"worth_it": "值得做", "maybe": "待定", "not_worth_it": "不值得", "error": "错误"}
VERDICT_COLORS = {"worth_it": "27AE60", "maybe": "F39C12", "not_worth_it": "E74C3C", "error": "95A5A6"}
SCORE_COLORS = {
    10: "1ABC9C", 9: "2ECC71", 8: "27AE60", 7: "F1C40F",
    6: "F39C12", 5: "E67E22", 4: "E74C3C", 3: "C0392B",
    2: "95A5A6", 1: "7F8C8D", 0: "BDC3C7",
}

COLS = [
    ("标题(中文)",     35),
    ("标题(原文)",     30),
    ("作者",           12),
    ("标签",           15),
    ("内容摘要(中文)", 45),
    ("社区回复(中文)", 30),
    ("回复数",         6),
    ("模块",           10),
    ("AI评分",         6),
    ("用户价值",       6),
    ("商业影响",       6),
    ("可行性",         6),
    ("结论",           8),
    ("评分理由",       40),
    ("创建日期",       12),
    ("是否最近7天",    10),
]


def _write_sheet(ws, records, title_text: str):
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"))

    for ci, (label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.freeze_panes = "A2"

    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

    for ri, r in enumerate(records, 2):
        is_recent = r.get("created_at", "") >= cutoff
        created_date = r.get("created_at", "")[:10]
        verdict = r.get("verdict", "")
        score = r.get("overall_score", 0)

        row_data = [
            r.get("title_zh") or r.get("title", ""),
            r.get("title", ""),
            r.get("author", ""),
            ", ".join(r.get("tags", [])) or "—",
            r.get("content_zh") or r.get("content", "")[:200],
            r.get("replies_zh") or "—",
            r.get("reply_count", 0),
            r.get("component", ""),
            score,
            r.get("user_value", 0),
            r.get("business_impact", 0),
            r.get("feasibility", 0),
            VERDICT_ZH.get(verdict, verdict),
            r.get("reason_zh", ""),
            created_date,
            "✓ 新" if is_recent else "",
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = thin_border

        # Color: score column
        score_cell = ws.cell(row=ri, column=9)
        sc = int(score) if score else 0
        if sc in SCORE_COLORS:
            score_cell.fill = PatternFill(
                start_color=SCORE_COLORS[sc],
                end_color=SCORE_COLORS[sc], fill_type="solid")
            if sc >= 7:
                score_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        # Color: verdict column
        v_cell = ws.cell(row=ri, column=13)
        if verdict in VERDICT_COLORS:
            v_cell.fill = PatternFill(
                start_color=VERDICT_COLORS[verdict],
                end_color=VERDICT_COLORS[verdict], fill_type="solid")
            v_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        # Highlight recent
        if is_recent:
            for ci in range(1, len(COLS) + 1):
                c = ws.cell(row=ri, column=ci)
                if ci not in (9, 13):
                    c.fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")


def do_export():
    if openpyxl is None:
        print("❌ pip install openpyxl"); return

    if TRANSLATED_PATH.is_file():
        records = json.load(open(TRANSLATED_PATH, encoding="utf-8"))
    else:
        print("⚠️  未翻译，使用原始数据（先运行 translate）")
        records = load_data()

    records.sort(key=lambda r: r.get("overall_score", 0), reverse=True)

    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    recent = [r for r in records if r.get("created_at", "") >= cutoff]

    wb = openpyxl.Workbook()

    # Sheet 1: 最近7天
    ws1 = wb.active
    ws1.title = "最近7天"
    if recent:
        recent.sort(key=lambda r: r.get("created_at", ""), reverse=True)
        _write_sheet(ws1, recent, "最近7天 Feature Request")
    else:
        ws1.cell(row=1, column=1, value="最近7天无新增 Feature Request")

    # Sheet 2: 全部（按分数排序）
    ws2 = wb.create_sheet("全部(按评分)")
    _write_sheet(ws2, records, "全部 Feature Request")

    # Sheet 3: 统计
    ws3 = wb.create_sheet("统计")
    sf = Font(name="Arial", size=11)
    bf = Font(name="Arial", size=13, bold=True)

    ws3.cell(row=1, column=1, value="Feature Request 看板统计").font = bf
    ws3.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws3.cell(row=3, column=1, value=f"总计: {len(records)} 条").font = sf
    ws3.cell(row=4, column=1, value=f"最近7天新增: {len(recent)} 条").font = sf

    valid = [r for r in records if r.get("overall_score", 0) > 0]
    if valid:
        avg = sum(r["overall_score"] for r in valid) / len(valid)
        worth = sum(1 for r in valid if r.get("verdict") == "worth_it")
        maybe = sum(1 for r in valid if r.get("verdict") == "maybe")
        nw = sum(1 for r in valid if r.get("verdict") == "not_worth_it")

        ws3.cell(row=5, column=1, value=f"平均分: {avg:.1f}/10").font = sf
        ws3.cell(row=6, column=1, value=f"值得做: {worth} | 待定: {maybe} | 不值得: {nw}").font = sf

        ws3.cell(row=8, column=1, value="TOP 10 最值得做").font = bf
        ws3.cell(row=9, column=1, value="评分").font = Font(bold=True)
        ws3.cell(row=9, column=2, value="标题").font = Font(bold=True)
        ws3.cell(row=9, column=3, value="模块").font = Font(bold=True)
        ws3.cell(row=9, column=4, value="理由").font = Font(bold=True)

        top10 = sorted(valid, key=lambda x: x["overall_score"], reverse=True)[:10]
        for i, r in enumerate(top10, 10):
            ws3.cell(row=i, column=1, value=f"{r['overall_score']}/10")
            ws3.cell(row=i, column=2, value=r.get("title_zh") or r["title"])
            ws3.cell(row=i, column=3, value=r.get("component", ""))
            ws3.cell(row=i, column=4, value=(r.get("reason_zh") or "")[:80])

        # By component
        row_off = 10 + len(top10) + 2
        ws3.cell(row=row_off, column=1, value="按模块统计").font = bf
        by_comp = {}
        for r in records:
            comp = r.get("component") or "未分类"
            by_comp.setdefault(comp, []).append(r)
        for j, (comp, items) in enumerate(sorted(by_comp.items(),
                                                   key=lambda x: len(x[1]), reverse=True)):
            avg_s = sum(r.get("overall_score", 0) for r in items) / max(len(items), 1)
            ws3.cell(row=row_off + 1 + j, column=1, value=comp)
            ws3.cell(row=row_off + 1 + j, column=2, value=f"{len(items)} 条")
            ws3.cell(row=row_off + 1 + j, column=3, value=f"均分 {avg_s:.1f}")

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 50

    wb.save(EXCEL_PATH)
    print(f"✅ 导出完成 → {EXCEL_PATH}")
    print(f"   全部: {len(records)} 条 | 最近7天: {len(recent)} 条")
    subprocess.run(["open", str(EXCEL_PATH)], check=False)


# ═══════════════════════════════════════════════════════════════════
#  SYNC — Feishu
# ═══════════════════════════════════════════════════════════════════

def fs_token(cfg: dict) -> str:
    global _fs_token, _fs_token_exp
    now = time.time()
    if _fs_token and now < _fs_token_exp - 120:
        return _fs_token
    fs = cfg.get("feishu") or cfg
    r = _curl("POST", FEISHU_TOKEN_URL,
              {"Content-Type": "application/json"},
              {"app_id": fs["app_id"], "app_secret": fs["app_secret"]})
    if isinstance(r, dict) and r.get("code") != 0:
        raise RuntimeError(f"飞书 token: {r}")
    _fs_token = r["tenant_access_token"]
    _fs_token_exp = now + int(r.get("expire", 7200))
    return _fs_token


def fs_create(token: str, app_token: str, table_id: str, fields: dict):
    url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records"
    r = _curl("POST", url,
              {"Authorization": f"Bearer {token}",
               "Content-Type": "application/json"},
              {"fields": fields})
    if isinstance(r, dict) and r.get("code") != 0:
        raise RuntimeError(f"飞书写入: {r}")


def do_sync(cfg: dict, table_id: str):
    fs = cfg.get("feishu") or cfg
    app_token = fs.get("app_token", "")
    token = fs_token(cfg)

    if TRANSLATED_PATH.is_file():
        records = json.load(open(TRANSLATED_PATH, encoding="utf-8"))
    else:
        records = load_data()

    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

    print(f"📤 推送 {len(records)} 条到飞书 (table: {table_id}) ...")
    ok = 0
    for r in records:
        is_recent = r.get("created_at", "") >= cutoff
        fields = {
            "标题": r.get("title_zh") or r.get("title", ""),
            "原标题": r.get("title", ""),
            "作者": r.get("author", ""),
            "标签": ", ".join(r.get("tags", [])) or "—",
            "内容摘要": r.get("content_zh") or r.get("content", "")[:200],
            "社区回复": r.get("replies_zh") or "—",
            "回复数": str(r.get("reply_count", 0)),
            "模块": r.get("component", ""),
            "AI评分": f"{r.get('overall_score', 0)}/10",
            "用户价值": str(r.get("user_value", 0)),
            "商业影响": str(r.get("business_impact", 0)),
            "可行性": str(r.get("feasibility", 0)),
            "结论": VERDICT_ZH.get(r.get("verdict", ""), r.get("verdict", "")),
            "评分理由": r.get("reason_zh", ""),
            "创建日期": r.get("created_at", "")[:10],
            "最近一周": "✓" if is_recent else "",
        }
        try:
            fs_create(token, app_token, table_id, fields)
            ok += 1
        except Exception as e:
            print(f"  ⚠️ {r.get('title','')[:30]}: {e}")
        time.sleep(0.3)
    print(f"✅ 推送完成: {ok}/{len(records)}")


# ═══════════════════════════════════════════════════════════════════
#  RUN
# ═══════════════════════════════════════════════════════════════════

def do_run(api_key: str, model: str):
    print("🚀 一键生成 Feature Request 看板\n")
    print("=" * 50)
    print("STEP 1/2: 翻译")
    print("=" * 50)
    do_translate(api_key, model)
    print(f"\n{'=' * 50}")
    print("STEP 2/2: 导出 Excel")
    print("=" * 50)
    do_export()


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description="Feature Request 看板 (中文翻译 + AI评分)")
    sub = p.add_subparsers(dest="cmd")

    sub.add_parser("translate", help="LLM 翻译全部标题+内容为中文")
    sub.add_parser("export", help="导出 Excel (最近7天 + 全部 + 统计)")
    sub.add_parser("run", help="一键: translate + export")

    s = sub.add_parser("sync", help="推送到飞书多维表格")
    s.add_argument("--table-id", required=True)

    args = p.parse_args()
    if not args.cmd:
        p.print_help(); return

    cfg = load_config()
    api_key = (cfg.get("openrouter_api_key") or "").strip()
    model = cfg.get("openrouter_model", "openai/gpt-4o-mini")

    if args.cmd == "translate":
        if not api_key:
            print("❌ openrouter_api_key 未配置"); return
        do_translate(api_key, model)
    elif args.cmd == "export":
        do_export()
    elif args.cmd == "run":
        if not api_key:
            print("❌ openrouter_api_key 未配置"); return
        do_run(api_key, model)
    elif args.cmd == "sync":
        do_sync(cfg, args.table_id)


if __name__ == "__main__":
    main()
