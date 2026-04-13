#!/usr/bin/env python3
"""
Telegram Channel/Group Scraper — Telethon

批量爬取 TG 公开频道/群的消息，导出 JSONL + 可选推飞书。
私有群（+邀请链接）需先手动加入。

用法:
  python3 tg_channel_scraper.py scrape --file channels.txt --limit 100
  python3 tg_channel_scraper.py scrape --channels LonesomeTheBlue_Official heniitrading --limit 50
  python3 tg_channel_scraper.py export             # → Excel
  python3 tg_channel_scraper.py sync --table-id tblXXX   # → 飞书
"""
from __future__ import annotations

import argparse
import asyncio
import json
import re
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    from telethon import TelegramClient
    from telethon.tl.functions.messages import GetHistoryRequest
    from telethon.tl.types import Channel, Chat, User, PeerChannel
    from telethon.errors import (
        ChannelPrivateError, ChannelInvalidError,
        FloodWaitError, ChatAdminRequiredError,
        InviteHashExpiredError, InviteHashInvalidError,
    )
    from telethon.tl.functions.messages import ImportChatInviteRequest, CheckChatInviteRequest
except ImportError:
    print("❌ pip install telethon"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

SCRIPT_DIR = Path(__file__).resolve().parent
TRACKER_DIR = SCRIPT_DIR.parent / "tg-partner-tracker"
SESSION_PATH = TRACKER_DIR / "tg_session"
CONFIG_PATH = TRACKER_DIR / "config.json"
DATA_DIR = SCRIPT_DIR / "tg_scrape_data"
EXCEL_PATH = SCRIPT_DIR / "tg_channels_export.xlsx"

FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_API = "https://open.feishu.cn/open-apis/bitable/v1"
FEISHU_CFG_PATH = SCRIPT_DIR.parent / "astrbot_plugin_dc_user_collector" / "feishu_config.json"

_fs_token: str | None = None
_fs_token_exp: float = 0.0


def load_tracker_config() -> dict:
    if CONFIG_PATH.is_file():
        return json.loads(CONFIG_PATH.read_text("utf-8"))
    print(f"❌ 找不到 {CONFIG_PATH}"); sys.exit(1)


def parse_channel_input(raw: str) -> dict:
    """Parse a TG link into structured info."""
    raw = raw.strip()
    if not raw or raw == "https://t.me/+":
        return {"type": "invalid", "raw": raw}

    # Message link (e.g. .../masteranandatrades/8301)
    m = re.match(r'https?://(?:www\.)?t\.me/(\w+)/(\d+)$', raw)
    if m:
        return {"type": "message_link", "username": m.group(1), "msg_id": m.group(2), "raw": raw}

    # Bot link
    if "?start=" in raw or raw.endswith("_bot"):
        m2 = re.match(r'https?://(?:www\.)?t\.me/(\w+)', raw)
        return {"type": "bot", "username": m2.group(1) if m2 else raw, "raw": raw}

    # Private invite link
    m = re.match(r'https?://(?:www\.)?t\.me/\+(\S+)', raw)
    if m:
        return {"type": "private", "hash": m.group(1), "raw": raw}

    # Public channel/group
    m = re.match(r'https?://(?:www\.)?t\.me/(\w+)', raw)
    if m:
        return {"type": "public", "username": m.group(1), "raw": raw}

    # Plain username
    if re.match(r'^\w+$', raw):
        return {"type": "public", "username": raw, "raw": raw}

    return {"type": "unknown", "raw": raw}


def parse_links_list(links: list[str]) -> list[dict]:
    results = []
    for link in links:
        link = link.strip()
        if not link:
            continue
        info = parse_channel_input(link)
        results.append(info)
    return results


# ═══════════════════════════════════════════════════════════════════
#  SCRAPE
# ═══════════════════════════════════════════════════════════════════

async def scrape_channel(client: TelegramClient, info: dict,
                         limit: int) -> list[dict]:
    """Scrape messages from a single channel/group."""
    ch_type = info["type"]

    if ch_type == "invalid":
        print(f"  ⏭ 无效链接: {info['raw']}")
        return []
    if ch_type == "bot":
        print(f"  ⏭ Bot 链接跳过: {info['raw']}")
        return []
    if ch_type == "message_link":
        print(f"  ⏭ 消息链接跳过: {info['raw']}")
        return []

    try:
        if ch_type == "public":
            entity = await client.get_entity(info["username"])
        elif ch_type == "private":
            try:
                check = await client(CheckChatInviteRequest(info["hash"]))
                if hasattr(check, 'chat'):
                    entity = check.chat
                else:
                    await client(ImportChatInviteRequest(info["hash"]))
                    await asyncio.sleep(2)
                    entity = await client.get_entity(info["hash"])
            except InviteHashExpiredError:
                print(f"  ❌ 邀请链接已过期: {info['raw']}")
                return []
            except InviteHashInvalidError:
                print(f"  ❌ 邀请链接无效: {info['raw']}")
                return []
            except Exception as e:
                if "already" in str(e).lower() or "INVITE_REQUEST_SENT" in str(e):
                    print(f"  ⏳ 需要管理员审批: {info['raw']}")
                else:
                    print(f"  ❌ 无法加入: {info['raw']} ({e})")
                return []
        else:
            print(f"  ⏭ 未知类型: {info['raw']}")
            return []

        ch_name = getattr(entity, 'title', '') or getattr(entity, 'username', '') or str(entity.id)
        ch_id = str(entity.id)

    except ChannelPrivateError:
        print(f"  🔒 私有频道无权限: {info['raw']}")
        return []
    except (ChannelInvalidError, ValueError) as e:
        print(f"  ❌ 频道不存在: {info['raw']} ({e})")
        return []
    except FloodWaitError as e:
        print(f"  ⏳ Flood wait {e.seconds}s, 等待...")
        await asyncio.sleep(e.seconds + 1)
        return await scrape_channel(client, info, limit)
    except Exception as e:
        print(f"  ❌ 获取频道失败: {info['raw']} ({e})")
        return []

    print(f"  📥 {ch_name} (id={ch_id})...", end=" ", flush=True)

    messages = []
    try:
        async for msg in client.iter_messages(entity, limit=limit):
            if msg.message or msg.media:
                sender_name = ""
                if msg.sender:
                    sender_name = (getattr(msg.sender, 'first_name', '') or '') + \
                                  (' ' + (getattr(msg.sender, 'last_name', '') or '')).rstrip()
                    if not sender_name.strip():
                        sender_name = getattr(msg.sender, 'title', '') or \
                                      getattr(msg.sender, 'username', '') or ''

                ts = msg.date.strftime("%Y-%m-%d %H:%M:%S UTC") if msg.date else ""

                rec = {
                    "msg_id": str(msg.id),
                    "channel_id": ch_id,
                    "channel_name": ch_name,
                    "channel_link": info.get("raw", ""),
                    "sender_name": sender_name.strip(),
                    "text": msg.message or "",
                    "has_media": bool(msg.media),
                    "views": getattr(msg, 'views', 0) or 0,
                    "forwards": getattr(msg, 'forwards', 0) or 0,
                    "timestamp": ts,
                }
                messages.append(rec)
    except ChatAdminRequiredError:
        print(f"需要管理员权限")
        return []
    except Exception as e:
        print(f"读取失败: {e}")
        return []

    print(f"{len(messages)} 条")
    return messages


async def do_scrape(links: list[str], limit: int):
    cfg = load_tracker_config()
    api_id = int(cfg["api_id"])
    api_hash = cfg["api_hash"]

    parsed = parse_links_list(links)

    # Stats
    public = [p for p in parsed if p["type"] == "public"]
    private = [p for p in parsed if p["type"] == "private"]
    invalid = [p for p in parsed if p["type"] in ("invalid", "bot", "message_link", "unknown")]

    print(f"📊 解析链接: {len(parsed)} 条")
    print(f"   公开频道: {len(public)}")
    print(f"   私有群(邀请): {len(private)}")
    print(f"   跳过(无效/bot): {len(invalid)}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    client = TelegramClient(str(SESSION_PATH), api_id, api_hash)
    await client.start()

    all_records = []
    success = 0
    failed = 0

    valid = [p for p in parsed if p["type"] in ("public", "private")]
    for i, info in enumerate(valid):
        print(f"\n[{i+1}/{len(valid)}]", end="")
        try:
            records = await scrape_channel(client, info, limit)
            if records:
                all_records.extend(records)
                success += 1
                for rec in records:
                    _append_jsonl(DATA_DIR / "channels.jsonl", rec)
            else:
                failed += 1
        except FloodWaitError as e:
            print(f"  ⏳ Flood wait {e.seconds}s...")
            await asyncio.sleep(e.seconds + 1)
            failed += 1

        await asyncio.sleep(1.5)

    # Also handle invalid ones
    for info in invalid:
        print(f"\n  ⏭ 跳过: {info['raw'][:60]}")

    await client.disconnect()

    print(f"\n{'═' * 50}")
    print(f"✅ 爬取完成")
    print(f"   成功: {success} 个频道/群")
    print(f"   失败: {failed} 个")
    print(f"   总消息: {len(all_records)} 条")
    print(f"   数据: {DATA_DIR / 'channels.jsonl'}")

    return all_records


def _append_jsonl(path: Path, rec: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def _read_jsonl(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    out = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    out.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return out


# ═══════════════════════════════════════════════════════════════════
#  EXPORT — Excel
# ═══════════════════════════════════════════════════════════════════

def do_export():
    if openpyxl is None:
        print("❌ pip install openpyxl"); return

    records = _read_jsonl(DATA_DIR / "channels.jsonl")
    if not records:
        print("❌ 没有数据，先 scrape"); return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TG Channels"

    cols = [
        ("频道", 25), ("发送者", 15), ("消息内容", 60),
        ("浏览数", 8), ("转发数", 8), ("时间", 20), ("频道链接", 40),
    ]

    hf = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    for ci, (label, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = hf
        cell.font = hfont
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"

    for ri, r in enumerate(records, 2):
        row = [
            r.get("channel_name", ""),
            r.get("sender_name", ""),
            r.get("text", "")[:500],
            r.get("views", 0),
            r.get("forwards", 0),
            r.get("timestamp", ""),
            r.get("channel_link", ""),
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Summary sheet
    ws2 = wb.create_sheet("统计")
    by_ch = {}
    for r in records:
        ch = r.get("channel_name", "?")
        by_ch.setdefault(ch, []).append(r)

    ws2.cell(row=1, column=1, value="频道统计").font = Font(size=14, bold=True)
    ws2.cell(row=2, column=1, value="频道").font = Font(bold=True)
    ws2.cell(row=2, column=2, value="消息数").font = Font(bold=True)
    ws2.cell(row=2, column=3, value="总浏览").font = Font(bold=True)

    for i, (ch, msgs) in enumerate(sorted(by_ch.items(), key=lambda x: len(x[1]), reverse=True), 3):
        ws2.cell(row=i, column=1, value=ch)
        ws2.cell(row=i, column=2, value=len(msgs))
        ws2.cell(row=i, column=3, value=sum(m.get("views", 0) for m in msgs))

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12

    wb.save(EXCEL_PATH)
    print(f"✅ 导出 {len(records)} 条 → {EXCEL_PATH}")
    subprocess.run(["open", str(EXCEL_PATH)], check=False)


# ═══════════════════════════════════════════════════════════════════
#  SYNC — Feishu
# ═══════════════════════════════════════════════════════════════════

def _curl(
    method: str,
    url: str,
    headers: dict,
    body: dict | None = None,
    *,
    timeout: int = 30,
) -> dict:
    cmd = ["curl", "-s", "-X", method, url]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    if body:
        cmd += ["-d", json.dumps(body)]
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    return json.loads(p.stdout) if p.stdout.strip() else {}


def fs_token() -> str:
    global _fs_token, _fs_token_exp
    now = time.time()
    if _fs_token and now < _fs_token_exp - 120:
        return _fs_token
    cfg = json.loads(FEISHU_CFG_PATH.read_text("utf-8"))
    r = _curl("POST", FEISHU_TOKEN_URL,
              {"Content-Type": "application/json"},
              {"app_id": cfg["app_id"], "app_secret": cfg["app_secret"]})
    _fs_token = r["tenant_access_token"]
    _fs_token_exp = now + int(r.get("expire", 7200))
    return _fs_token


def _record_to_fields(r: dict, *, link_as_url: bool = True) -> dict:
    link = (r.get("channel_link") or "").strip()
    if link_as_url and link:
        link_val: str | dict = {"link": link, "text": "链接"}
    else:
        link_val = link
    return {
        "频道": r.get("channel_name", ""),
        "发送者": r.get("sender_name", ""),
        "消息内容": (r.get("text") or "")[:2000],
        "浏览数": str(r.get("views", 0)),
        "转发数": str(r.get("forwards", 0)),
        "时间": r.get("timestamp", ""),
        "频道链接": link_val,
    }


def do_sync(table_id: str):
    records = _read_jsonl(DATA_DIR / "channels.jsonl")
    if not records:
        print("❌ 没有数据"); return

    cfg = json.loads(FEISHU_CFG_PATH.read_text("utf-8"))
    app_token = cfg["app_token"]
    token = fs_token()

    batch_url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records/batch_create"
    single_url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records"
    batch_size = 100
    total = len(records)
    ok = 0
    use_url_field = True

    print(f"📤 批量推送 {total} 条到飞书 (每批最多 {batch_size})...")

    for i in range(0, total, batch_size):
        chunk = records[i : i + batch_size]
        body = {
            "records": [
                {"fields": _record_to_fields(r, link_as_url=use_url_field)}
                for r in chunk
            ],
        }
        resp = _curl(
            "POST",
            batch_url,
            {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            body,
            timeout=120,
        )
        if resp.get("code") == 0:
            ok += len(chunk)
            print(f"  [{min(i + batch_size, total)}/{total}] ✓")
        else:
            msg = str(resp.get("msg", ""))
            if use_url_field and ("频道链接" in msg or "FieldName" in msg or "1254045" in str(resp)):
                use_url_field = False
                body = {
                    "records": [
                        {"fields": _record_to_fields(r, link_as_url=False)}
                        for r in chunk
                    ],
                }
                resp = _curl(
                    "POST",
                    batch_url,
                    {
                        "Authorization": f"Bearer {token}",
                        "Content-Type": "application/json",
                    },
                    body,
                    timeout=120,
                )
            if resp.get("code") == 0:
                ok += len(chunk)
                print(f"  [{min(i + batch_size, total)}/{total}] ✓ (文本链接)")
            else:
                print(f"  ⚠️ 批次 {i // batch_size + 1} 失败，改单条: {resp.get('msg', resp)[:120]}")
                for r in chunk:
                    one = _curl(
                        "POST",
                        single_url,
                        {
                            "Authorization": f"Bearer {token}",
                            "Content-Type": "application/json",
                        },
                        {"fields": _record_to_fields(r, link_as_url=use_url_field)},
                        timeout=60,
                    )
                    if one.get("code") != 0 and use_url_field:
                        one = _curl(
                            "POST",
                            single_url,
                            {
                                "Authorization": f"Bearer {token}",
                                "Content-Type": "application/json",
                            },
                            {"fields": _record_to_fields(r, link_as_url=False)},
                            timeout=60,
                        )
                    if one.get("code") == 0:
                        ok += 1
                    time.sleep(0.05)
        time.sleep(0.35)

    print(f"✅ 完成: {ok}/{total}")


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description="TG Channel Scraper (Telethon)")
    sub = p.add_subparsers(dest="cmd")

    s = sub.add_parser("scrape", help="爬取频道/群消息")
    s.add_argument("--channels", nargs="+", default=[], help="频道用户名或链接")
    s.add_argument("--file", type=str, default="", help="包含链接的文本文件")
    s.add_argument("--limit", type=int, default=100, help="每个频道最多拉取条数")
    s.add_argument("--clean", action="store_true", help="清空旧数据后再爬")

    sub.add_parser("export", help="导出 Excel")

    sy = sub.add_parser("sync", help="推送飞书")
    sy.add_argument("--table-id", required=True)

    args = p.parse_args()
    if not args.cmd:
        p.print_help(); return

    if args.cmd == "scrape":
        links = list(args.channels)
        if args.file:
            fp = Path(args.file)
            if fp.is_file():
                links.extend(fp.read_text("utf-8").strip().splitlines())
            else:
                print(f"❌ 文件不存在: {fp}"); return
        if not links:
            print("❌ 请提供 --channels 或 --file"); return
        if args.clean and (DATA_DIR / "channels.jsonl").is_file():
            (DATA_DIR / "channels.jsonl").unlink()
            print("🗑 已清空旧数据")
        asyncio.run(do_scrape(links, args.limit))
    elif args.cmd == "export":
        do_export()
    elif args.cmd == "sync":
        do_sync(args.table_id)


if __name__ == "__main__":
    main()
