#!/usr/bin/env python3
"""
Discord Bug Pipeline — Minara Support 频道全流程

1. fetch   — 从 bug-report / feedback 频道 + ticket-* 中 @hazel 的消息拉取到本地
2. ack     — 给每条消息回复确认 "Got it, thanks for reporting!"
3. analyze — LLM 分析每条 bug/feedback → 写入本地 Excel
4. export  — 导出 Excel 供人工审核
5. sync    — 手动推到飞书多维表格

用法:
  python3 dc_bug_pipeline.py fetch                        # 拉取
  python3 dc_bug_pipeline.py fetch --guild-id 123456      # 指定服务器
  python3 dc_bug_pipeline.py ack                          # 发送确认回复
  python3 dc_bug_pipeline.py ack --dry-run                # 预览不发送
  python3 dc_bug_pipeline.py analyze                      # LLM 分析
  python3 dc_bug_pipeline.py export                       # 导出 Excel
  python3 dc_bug_pipeline.py sync                         # 推飞书
  python3 dc_bug_pipeline.py run                          # 一键: fetch → ack → analyze → export
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
DATA_DIR = SCRIPT_DIR / "bug_pipeline_data"
EXCEL_PATH = SCRIPT_DIR / "bug_report.xlsx"

CONFIG_CANDIDATES = [
    BASE_DIR / "astrbot_plugin_dc_assistant" / "config.json",
    BASE_DIR / "astrbot_plugin_dc_user_collector" / "feishu_config.json",
]

DISCORD_API = "https://discord.com/api/v10"
FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_API = "https://open.feishu.cn/open-apis/bitable/v1"
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

SUPPORT_KEYWORDS = {"bug-report", "bug-reports", "feedback"}
TICKET_KEYWORDS = {"ticket-", "error-", "login-", "deposit", "hacked", "refund", "regain"}
HAZEL_NAMES = {"hazel"}

ACK_MESSAGE = (
    "Got it, thanks for reporting! Our team is looking into this "
    "and will get back to you shortly."
)

BUG_ANALYSIS_PROMPT = """\
You are a senior QA engineer at Minara, a crypto trading platform with AI copilot, \
autopilot trading, workflows, and smart wallet features.

Analyze this bug report / user feedback and output a JSON object (no markdown wrapping):

{{
  "category": "<bug | ux_issue | feature_request | question | praise | other>",
  "severity": "<critical | high | medium | low>",
  "component": "<wallet | trading | workflow | autopilot | copilot | ui | auth | other>",
  "summary_en": "<1-sentence English summary, max 80 chars>",
  "summary_zh": "<1-sentence Chinese summary, max 50 chars>",
  "reproducible": "<yes | no | unclear>",
  "affected_platform": "<web | mobile | api | unknown>",
  "suggested_action": "<1-sentence what the dev team should do>",
  "user_sentiment": "<frustrated | neutral | constructive | positive>"
}}

Channel: #{channel}
Author: {author}
Message:
{content}

Thread context (if any):
{context}
"""

_fs_token: str | None = None
_fs_token_exp: float = 0.0


# ═══════════════════════════════════════════════════════════════════
#  Config & HTTP
# ═══════════════════════════════════════════════════════════════════

def load_config() -> dict:
    for p in CONFIG_CANDIDATES:
        if p.is_file():
            cfg = json.loads(p.read_text("utf-8"))
            if cfg.get("discord_bot_token"):
                return cfg
            if cfg.get("feishu"):
                return cfg
    print("❌ 找不到 config.json"); sys.exit(1)


def bot_token(cfg: dict) -> str:
    return (cfg.get("discord_bot_token") or "").strip()


def llm_key(cfg: dict) -> str:
    return (cfg.get("openrouter_api_key") or "").strip()


def _curl(method: str, url: str, headers: dict,
          body: dict | None = None, timeout: int = 30) -> dict | list:
    cmd = ["curl", "-s", "-X", method, url]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    if body is not None:
        cmd += ["-d", json.dumps(body)]
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if not p.stdout.strip():
        return {}
    try:
        return json.loads(p.stdout)
    except json.JSONDecodeError:
        return {}


def dc_get(path: str, token: str) -> dict | list:
    return _curl("GET", f"{DISCORD_API}{path}",
                 {"Authorization": f"Bot {token}"})


def dc_post(path: str, token: str, body: dict) -> dict:
    r = _curl("POST", f"{DISCORD_API}{path}",
              {"Authorization": f"Bot {token}",
               "Content-Type": "application/json"}, body)
    return r if isinstance(r, dict) else {}


def append_jsonl(path: Path, rec: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def read_jsonl(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    out = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    out.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return out


def write_jsonl(path: Path, records: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ═══════════════════════════════════════════════════════════════════
#  FETCH — pull from Discord
# ═══════════════════════════════════════════════════════════════════

def discover_guild_id(token: str, guild_id: str | None) -> str:
    if guild_id:
        return guild_id
    guilds = dc_get("/users/@me/guilds", token)
    if not isinstance(guilds, list) or not guilds:
        print("❌ Bot 没有加入任何服务器"); sys.exit(1)
    if len(guilds) == 1:
        gid = guilds[0]["id"]
        print(f"  自动选择服务器: {guilds[0].get('name', gid)}")
        return gid
    print("  Bot 加入了多个服务器:")
    for g in guilds:
        print(f"    {g['id']} — {g.get('name', '?')}")
    print("  请用 --guild-id 指定"); sys.exit(1)


def _strip_emoji_prefix(name: str) -> str:
    """Strip emoji/separator prefixes like '🐞┃bug-report' → 'bug-report'."""
    for sep in ("┃", "│", "｜", "|", " "):
        if sep in name:
            name = name.split(sep, 1)[-1]
    return name.strip()


def find_support_channels(token: str, guild_id: str) -> dict[str, list[dict]]:
    """Find bug-report, feedback channels + ticket-like channels."""
    all_channels = dc_get(f"/guilds/{guild_id}/channels", token)
    if not isinstance(all_channels, list):
        print(f"❌ 获取频道列表失败: {all_channels}"); sys.exit(1)

    result = {"support": [], "tickets": []}

    for ch in all_channels:
        raw_name = (ch.get("name") or "").lower()
        clean = _strip_emoji_prefix(raw_name)
        ch_type = ch.get("type", 0)

        if clean in SUPPORT_KEYWORDS or any(kw in clean for kw in SUPPORT_KEYWORDS):
            result["support"].append(ch)
        elif any(clean.startswith(kw) or kw in clean for kw in TICKET_KEYWORDS):
            result["tickets"].append(ch)

    return result


def fetch_channel_messages(channel_id: str, token: str,
                           limit: int = 100) -> list[dict]:
    all_msgs = []
    before = None
    while len(all_msgs) < limit:
        batch = min(100, limit - len(all_msgs))
        path = f"/channels/{channel_id}/messages?limit={batch}"
        if before:
            path += f"&before={before}"
        msgs = dc_get(path, token)
        if not isinstance(msgs, list) or not msgs:
            break
        all_msgs.extend(msgs)
        if len(msgs) < batch:
            break
        before = msgs[-1]["id"]
        time.sleep(0.5)
    all_msgs.sort(key=lambda m: m["id"])
    return all_msgs


def mentions_hazel(msg: dict) -> bool:
    """Check if a message @mentions hazel."""
    content = (msg.get("content") or "").lower()
    for name in HAZEL_NAMES:
        if f"@{name}" in content or name in content:
            return True
    mentions = msg.get("mentions") or []
    for m in mentions:
        uname = (m.get("username") or "").lower()
        gname = (m.get("global_name") or "").lower()
        if uname in HAZEL_NAMES or gname in HAZEL_NAMES:
            return True
    return False


def msg_to_record(msg: dict, channel_name: str, channel_id: str,
                  source: str) -> dict:
    author = msg.get("author", {})
    ts = msg.get("timestamp") or ""
    if ts:
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            ts = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
        except ValueError:
            pass
    return {
        "msg_id": msg.get("id", ""),
        "channel_id": channel_id,
        "channel_name": channel_name,
        "source": source,
        "author_id": author.get("id", ""),
        "author_name": author.get("global_name") or author.get("username") or "",
        "author_username": author.get("username", ""),
        "is_bot": author.get("bot", False),
        "content": msg.get("content", ""),
        "timestamp": ts,
        "mentions_hazel": mentions_hazel(msg),
        "ack_sent": False,
        "analysis": None,
    }


def _fetch_forum_threads(channel_id: str, token: str) -> list[dict]:
    """Fetch active + archived threads from a forum channel."""
    all_threads = []
    ch_info = dc_get(f"/channels/{channel_id}", token)
    guild_id = ch_info.get("guild_id") if isinstance(ch_info, dict) else None
    if guild_id:
        active = dc_get(f"/guilds/{guild_id}/threads/active", token)
        if isinstance(active, dict):
            for t in active.get("threads") or []:
                if str(t.get("parent_id")) == str(channel_id):
                    all_threads.append(t)

    before = None
    while True:
        path = f"/channels/{channel_id}/threads/archived/public?limit=100"
        if before:
            path += f"&before={before}"
        data = dc_get(path, token)
        if not isinstance(data, dict):
            break
        threads = data.get("threads") or []
        if not threads:
            break
        all_threads.extend(threads)
        if not data.get("has_more"):
            break
        before = threads[-1].get("thread_metadata", {}).get("archive_timestamp")
        time.sleep(0.5)
    return all_threads


def do_fetch(cfg: dict, guild_id: str | None, limit: int):
    token = bot_token(cfg)
    if not token:
        print("❌ discord_bot_token 未配置"); return

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    gid = discover_guild_id(token, guild_id)

    print("\n=== 发现频道 ===")
    channels = find_support_channels(token, gid)

    print(f"  Support 频道: {[c['name'] for c in channels['support']]}")
    print(f"  Ticket 频道: {len(channels['tickets'])} 个")

    existing = read_jsonl(DATA_DIR / "bugs.jsonl")
    seen_ids = {r["msg_id"] for r in existing}
    new_records = []

    # 1) Pull from support channels (bug-report, feedback) — ALL messages
    for ch in channels["support"]:
        ch_id = ch["id"]
        ch_name = ch["name"]
        ch_type = ch.get("type", 0)

        if ch_type == 15:
            # Forum channel — fetch threads then messages from each thread
            print(f"\n📥 拉取 #{ch_name} (论坛, 遍历帖子)...")
            threads = _fetch_forum_threads(ch_id, token)
            count = 0
            for thread in threads:
                tid = thread["id"]
                t_name = thread.get("name", tid)
                msgs = fetch_channel_messages(tid, token, limit=limit)
                for msg in msgs:
                    if msg["id"] in seen_ids:
                        continue
                    if msg.get("author", {}).get("bot"):
                        continue
                    rec = msg_to_record(msg, f"{ch_name}/{t_name}", ch_id, source="support")
                    rec["thread_id"] = tid
                    rec["thread_title"] = t_name
                    new_records.append(rec)
                    seen_ids.add(msg["id"])
                    count += 1
                time.sleep(0.3)
            print(f"  {len(threads)} 个帖子, 新增 {count} 条消息")
        else:
            print(f"\n📥 拉取 #{ch_name} (全部消息)...")
            msgs = fetch_channel_messages(ch_id, token, limit=limit)
            count = 0
            for msg in msgs:
                if msg["id"] in seen_ids:
                    continue
                if msg.get("author", {}).get("bot"):
                    continue
                rec = msg_to_record(msg, ch_name, ch_id, source="support")
                new_records.append(rec)
                seen_ids.add(msg["id"])
                count += 1
            print(f"  新增 {count} 条 (共拉取 {len(msgs)})")

    # 2) Pull from ticket channels — only messages that @hazel
    ticket_with_hazel = 0
    total_tickets = len(channels["tickets"])

    if total_tickets > 0:
        print(f"\n📥 扫描 {total_tickets} 个 ticket 频道中 @Hazel 的消息...")

    for i, ch in enumerate(channels["tickets"]):
        ch_id = ch["id"]
        ch_name = ch["name"]

        msgs = fetch_channel_messages(ch_id, token, limit=50)

        hazel_msgs = [m for m in msgs if mentions_hazel(m)
                      and not m.get("author", {}).get("bot")
                      and m["id"] not in seen_ids]

        if hazel_msgs:
            ticket_with_hazel += 1
            ctx_texts = [f"[{m.get('author',{}).get('username','?')}]: {m.get('content','')[:200]}"
                         for m in msgs[:10]]

            for msg in hazel_msgs:
                rec = msg_to_record(msg, ch_name, ch_id, source="ticket")
                rec["thread_context"] = ctx_texts
                new_records.append(rec)
                seen_ids.add(msg["id"])

            print(f"  [{i+1}/{total_tickets}] #{ch_name}: {len(hazel_msgs)} 条 @Hazel ✓")

        if (i + 1) % 20 == 0:
            print(f"  ... 已扫描 {i+1}/{total_tickets} 个 ticket")
        time.sleep(0.3)

    # Save
    for rec in new_records:
        append_jsonl(DATA_DIR / "bugs.jsonl", rec)

    print(f"\n{'═' * 50}")
    print(f"✅ 新增 {len(new_records)} 条记录")
    print(f"   Support 频道: {sum(1 for r in new_records if r['source'] == 'support')}")
    print(f"   Ticket @Hazel: {sum(1 for r in new_records if r['source'] == 'ticket')}")
    print(f"   Ticket 频道含 @Hazel: {ticket_with_hazel}/{total_tickets}")
    print(f"   总计本地: {len(existing) + len(new_records)} 条")


# ═══════════════════════════════════════════════════════════════════
#  ACK — reply to each message
# ═══════════════════════════════════════════════════════════════════

def do_ack(cfg: dict, dry_run: bool = False):
    token = bot_token(cfg)
    records = read_jsonl(DATA_DIR / "bugs.jsonl")
    unacked = [r for r in records if not r.get("ack_sent")]

    if not unacked:
        print("✅ 所有消息都已确认回复"); return

    print(f"📤 发送确认回复: {len(unacked)} 条")
    sent = 0

    for r in unacked:
        ch_id = r["channel_id"]
        msg_id = r["msg_id"]
        author = r["author_name"] or r["author_username"]

        reply = ACK_MESSAGE
        if dry_run:
            print(f"  🔍 [#{r['channel_name']}] → {author}: {reply[:60]}")
            sent += 1
            continue

        body = {
            "content": reply,
            "message_reference": {"message_id": msg_id},
        }
        resp = dc_post(f"/channels/{ch_id}/messages", token, body)

        if resp.get("id"):
            r["ack_sent"] = True
            sent += 1
            print(f"  ✅ [#{r['channel_name']}] → {author}")
        else:
            print(f"  ❌ [#{r['channel_name']}] {resp}")
        time.sleep(0.5)

    if not dry_run:
        write_jsonl(DATA_DIR / "bugs.jsonl", records)

    tag = "🔍 DRY RUN" if dry_run else "📊"
    print(f"\n{tag} 已回复 {sent}/{len(unacked)} 条")


# ═══════════════════════════════════════════════════════════════════
#  ANALYZE — LLM bug analysis
# ═══════════════════════════════════════════════════════════════════

def llm_analyze_bug(api_key: str, model: str, rec: dict) -> dict:
    ctx = "\n".join(rec.get("thread_context") or [])[:1000]
    prompt = BUG_ANALYSIS_PROMPT.format(
        channel=rec.get("channel_name", "?"),
        author=rec.get("author_name", "?"),
        content=rec.get("content", "")[:2000],
        context=ctx or "(none)",
    )
    body = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
    }
    r = _curl("POST", OPENROUTER_URL,
              {"Authorization": f"Bearer {api_key}",
               "Content-Type": "application/json"}, body, timeout=60)
    text = (r.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
    return json.loads(text)


def do_analyze(cfg: dict):
    api_key = llm_key(cfg)
    if not api_key:
        print("❌ openrouter_api_key 未配置"); return
    model = cfg.get("openrouter_model", "openai/gpt-4o-mini")

    records = read_jsonl(DATA_DIR / "bugs.jsonl")
    unanalyzed = [r for r in records if not r.get("analysis")]

    if not unanalyzed:
        print("✅ 所有消息都已分析"); return

    print(f"🤖 LLM 分析: {len(unanalyzed)} 条\n")
    done = 0

    for r in unanalyzed:
        content_preview = (r.get("content") or "")[:60]
        print(f"  [{done+1}/{len(unanalyzed)}] #{r['channel_name']} | "
              f"{r['author_name']}: {content_preview}...", end=" ", flush=True)

        try:
            analysis = llm_analyze_bug(api_key, model, r)
            r["analysis"] = analysis
            done += 1
            sev = analysis.get("severity", "?")
            cat = analysis.get("category", "?")
            summary = analysis.get("summary_en", "?")
            print(f"→ [{sev}] {cat}: {summary[:40]}")
        except Exception as e:
            print(f"→ ❌ {e}")
            r["analysis"] = {
                "category": "error", "severity": "unknown",
                "summary_en": f"Analysis failed: {e}",
                "summary_zh": "分析失败",
            }
        time.sleep(1)

    write_jsonl(DATA_DIR / "bugs.jsonl", records)

    # Stats
    analyzed = [r for r in records if r.get("analysis") and r["analysis"].get("category") != "error"]
    if analyzed:
        by_sev = {}
        by_cat = {}
        for r in analyzed:
            a = r["analysis"]
            sev = a.get("severity", "?")
            cat = a.get("category", "?")
            by_sev[sev] = by_sev.get(sev, 0) + 1
            by_cat[cat] = by_cat.get(cat, 0) + 1
        print(f"\n{'═' * 50}")
        print(f"📊 分析摘要 ({len(analyzed)} 条)")
        print(f"  按严重度: {dict(sorted(by_sev.items(), key=lambda x: x[1], reverse=True))}")
        print(f"  按分类:   {dict(sorted(by_cat.items(), key=lambda x: x[1], reverse=True))}")


# ═══════════════════════════════════════════════════════════════════
#  EXPORT — local JSONL → Excel
# ═══════════════════════════════════════════════════════════════════

SEVERITY_COLORS = {
    "critical": "FF0000", "high": "FF6600",
    "medium": "FFCC00", "low": "66CC66",
}

COLS = [
    ("来源",         10),
    ("频道",         18),
    ("发送者",       14),
    ("消息内容",     55),
    ("时间",         20),
    ("分类",         14),
    ("严重度",       10),
    ("组件",         12),
    ("摘要(EN)",     40),
    ("摘要(ZH)",     30),
    ("建议操作",     40),
    ("可复现",       8),
    ("平台",         10),
    ("用户情绪",     10),
    ("已确认",       8),
    ("channel_id",   20),
    ("msg_id",       20),
]


def do_export():
    if openpyxl is None:
        print("❌ 需要 openpyxl:  pip install openpyxl"); return

    records = read_jsonl(DATA_DIR / "bugs.jsonl")
    if not records:
        print("❌ 没有数据，先运行 fetch"); return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bug Report"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    for ci, (label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.freeze_panes = "A2"

    for ri, r in enumerate(records, 2):
        a = r.get("analysis") or {}
        row_data = [
            r.get("source", ""),
            r.get("channel_name", ""),
            r.get("author_name", ""),
            r.get("content", ""),
            r.get("timestamp", ""),
            a.get("category", ""),
            a.get("severity", ""),
            a.get("component", ""),
            a.get("summary_en", ""),
            a.get("summary_zh", ""),
            a.get("suggested_action", ""),
            a.get("reproducible", ""),
            a.get("affected_platform", ""),
            a.get("user_sentiment", ""),
            "✓" if r.get("ack_sent") else "",
            r.get("channel_id", ""),
            r.get("msg_id", ""),
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = cell_font
            cell.alignment = wrap

        severity = a.get("severity", "")
        if severity in SEVERITY_COLORS:
            sev_cell = ws.cell(row=ri, column=7)
            sev_cell.fill = PatternFill(
                start_color=SEVERITY_COLORS[severity],
                end_color=SEVERITY_COLORS[severity],
                fill_type="solid")
            if severity in ("critical", "high"):
                sev_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    # Summary sheet
    ws2 = wb.create_sheet("统计")
    analyzed = [r for r in records if r.get("analysis")]

    ws2.cell(row=1, column=1, value="Bug Pipeline 统计").font = Font(size=14, bold=True)
    ws2.cell(row=2, column=1, value=f"总记录: {len(records)}")
    ws2.cell(row=3, column=1, value=f"已分析: {len(analyzed)}")
    ws2.cell(row=4, column=1, value=f"已确认: {sum(1 for r in records if r.get('ack_sent'))}")

    if analyzed:
        ws2.cell(row=6, column=1, value="按严重度").font = Font(bold=True)
        by_sev = {}
        for r in analyzed:
            s = r["analysis"].get("severity", "?")
            by_sev[s] = by_sev.get(s, 0) + 1
        for i, (k, v) in enumerate(sorted(by_sev.items()), 7):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=v)

        row_offset = 7 + len(by_sev) + 1
        ws2.cell(row=row_offset, column=1, value="按分类").font = Font(bold=True)
        by_cat = {}
        for r in analyzed:
            c = r["analysis"].get("category", "?")
            by_cat[c] = by_cat.get(c, 0) + 1
        for i, (k, v) in enumerate(sorted(by_cat.items()), row_offset + 1):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=v)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    wb.save(EXCEL_PATH)
    print(f"✅ 导出 {len(records)} 条 → {EXCEL_PATH}")
    subprocess.run(["open", str(EXCEL_PATH)], check=False)


# ═══════════════════════════════════════════════════════════════════
#  SYNC — push to Feishu Bitable
# ═══════════════════════════════════════════════════════════════════

def fs_token(cfg: dict) -> str:
    global _fs_token, _fs_token_exp
    now = time.time()
    if _fs_token and now < _fs_token_exp - 120:
        return _fs_token
    fs = cfg.get("feishu") or cfg
    r = _curl("POST", FEISHU_TOKEN_URL,
              {"Content-Type": "application/json"},
              {"app_id": fs["app_id"], "app_secret": fs["app_secret"]})
    if isinstance(r, dict) and r.get("code") != 0:
        raise RuntimeError(f"飞书 token: {r}")
    _fs_token = r["tenant_access_token"]
    _fs_token_exp = now + int(r.get("expire", 7200))
    return _fs_token


def fs_create(token: str, app_token: str, table_id: str, fields: dict):
    url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records"
    r = _curl("POST", url,
              {"Authorization": f"Bearer {token}",
               "Content-Type": "application/json"},
              {"fields": fields})
    if isinstance(r, dict) and r.get("code") != 0:
        raise RuntimeError(f"飞书写入: {r}")


def do_sync(cfg: dict, table_id: str | None = None):
    fs = cfg.get("feishu") or cfg
    if not fs.get("app_id"):
        print("❌ 飞书配置缺失"); return

    records = read_jsonl(DATA_DIR / "bugs.jsonl")
    analyzed = [r for r in records if r.get("analysis")]
    if not analyzed:
        print("❌ 没有已分析的数据，先运行 analyze"); return

    tid = table_id or fs.get("bug_table_id") or ""
    if not tid:
        print("❌ 请提供 --table-id 或在 config.json 中设置 feishu.bug_table_id")
        print("   飞书表需要以下字段: 来源, 频道, 发送者, 消息内容, 时间,")
        print("   分类, 严重度, 组件, 摘要, 建议操作")
        return

    app_token = fs.get("app_token", "")
    token = fs_token(cfg)

    print(f"📤 推送 {len(analyzed)} 条到飞书 (table: {tid}) ...")
    ok = 0
    for r in analyzed:
        a = r["analysis"]
        fields = {
            "来源": r.get("source", ""),
            "频道": r.get("channel_name", ""),
            "发送者": r.get("author_name", ""),
            "消息内容": (r.get("content") or "")[:2000],
            "时间": r.get("timestamp", ""),
            "分类": a.get("category", ""),
            "严重度": a.get("severity", ""),
            "组件": a.get("component", ""),
            "摘要": a.get("summary_zh", ""),
            "摘要EN": a.get("summary_en", ""),
            "建议操作": a.get("suggested_action", ""),
            "可复现": a.get("reproducible", ""),
            "平台": a.get("affected_platform", ""),
        }
        try:
            fs_create(token, app_token, tid, fields)
            ok += 1
        except Exception as e:
            print(f"  ⚠️ {e}")
        time.sleep(0.3)

    print(f"✅ 推送完成: {ok}/{len(analyzed)}")


# ═══════════════════════════════════════════════════════════════════
#  RUN — full pipeline
# ═══════════════════════════════════════════════════════════════════

def do_run(cfg: dict, guild_id: str | None, limit: int):
    print("🚀 一键执行完整 Pipeline\n")

    print("=" * 50)
    print("STEP 1/4: FETCH")
    print("=" * 50)
    do_fetch(cfg, guild_id, limit)

    print(f"\n{'=' * 50}")
    print("STEP 2/4: ACK (确认回复)")
    print("=" * 50)
    do_ack(cfg)

    print(f"\n{'=' * 50}")
    print("STEP 3/4: ANALYZE (LLM 分析)")
    print("=" * 50)
    do_analyze(cfg)

    print(f"\n{'=' * 50}")
    print("STEP 4/4: EXPORT (导出 Excel)")
    print("=" * 50)
    do_export()

    print(f"\n{'═' * 50}")
    print("✅ Pipeline 完成！")
    print(f"   Excel: {EXCEL_PATH}")
    print(f"   数据: {DATA_DIR / 'bugs.jsonl'}")
    print(f"   手动同步飞书: python3 dc_bug_pipeline.py sync --table-id tblXXX")


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(
        description="Discord Bug Pipeline — Minara Support 全流程")
    sub = p.add_subparsers(dest="cmd")

    f = sub.add_parser("fetch", help="从 Discord 拉取 bug-report/feedback + ticket @hazel")
    f.add_argument("--guild-id", type=str, default=None)
    f.add_argument("--limit", type=int, default=200)

    a = sub.add_parser("ack", help="发送确认回复")
    a.add_argument("--dry-run", action="store_true")

    sub.add_parser("analyze", help="LLM 分析每条 bug/feedback")
    sub.add_parser("export", help="导出 Excel")

    s = sub.add_parser("sync", help="推送到飞书多维表格")
    s.add_argument("--table-id", type=str, default=None)

    r = sub.add_parser("run", help="一键执行: fetch → ack → analyze → export")
    r.add_argument("--guild-id", type=str, default=None)
    r.add_argument("--limit", type=int, default=200)

    args = p.parse_args()
    if not args.cmd:
        p.print_help(); return

    cfg = load_config()

    if args.cmd == "fetch":
        do_fetch(cfg, args.guild_id, args.limit)
    elif args.cmd == "ack":
        do_ack(cfg, dry_run=args.dry_run)
    elif args.cmd == "analyze":
        do_analyze(cfg)
    elif args.cmd == "export":
        do_export()
    elif args.cmd == "sync":
        do_sync(cfg, table_id=args.table_id)
    elif args.cmd == "run":
        do_run(cfg, args.guild_id, args.limit)


if __name__ == "__main__":
    main()
