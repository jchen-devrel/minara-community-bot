#!/usr/bin/env python3
"""
Discord DevRel Reviewer — Local JSONL + Excel workflow

Flow:
  1. AstrBot dc_assistant plugin collects messages → data/*.jsonl
  2. python3 dc_review.py export   → mentions → review.xlsx (auto-opens)
  3. Edit Excel: A列(操作) = approve/edit/skip, I列(AI草稿) = modify text
  4. python3 dc_review.py send     → Excel → Discord batch reply
  5. python3 dc_review.py sync     → push all data to Feishu Bitable
  6. python3 dc_review.py fetch --from-config  → pull history (dedupe by msg_id)
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

SCRIPT_DIR = Path(__file__).resolve().parent
PLUGIN_DIR = SCRIPT_DIR.parent / "astrbot_plugin_dc_assistant"
CONFIG_PATH = SCRIPT_DIR / "config.json"
DATA_DIR = PLUGIN_DIR / "data"
EXCEL_PATH = SCRIPT_DIR / "review.xlsx"
REPLIES_LOG = SCRIPT_DIR / "replies.jsonl"

DISCORD_API = "https://discord.com/api/v10"
FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_API = "https://open.feishu.cn/open-apis/bitable/v1"
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

_fs_token: str | None = None
_fs_token_exp: float = 0.0


# ═══════════════════════════════════════════════════════════════════
#  Config
# ═══════════════════════════════════════════════════════════════════

def load_config() -> dict:
    if CONFIG_PATH.is_file():
        return json.loads(CONFIG_PATH.read_text("utf-8"))
    alt = PLUGIN_DIR / "config.json"
    if alt.is_file():
        return json.loads(alt.read_text("utf-8"))
    print("❌ 找不到 config.json"); sys.exit(1)


def read_jsonl(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    records = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return records


def append_jsonl(path: Path, rec: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ═══════════════════════════════════════════════════════════════════
#  HTTP (curl-based for Cloudflare compat)
# ═══════════════════════════════════════════════════════════════════

def _curl(method: str, url: str, headers: dict,
          body: dict | None = None, timeout: int = 30) -> dict | list:
    cmd = ["curl", "-s", "-X", method, url]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    if body is not None:
        cmd += ["-d", json.dumps(body)]
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if not p.stdout.strip():
        return {}
    try:
        return json.loads(p.stdout)
    except json.JSONDecodeError:
        print(f"  ⚠️ parse error: {p.stdout[:200]}", file=sys.stderr)
        return {}


def discord_get(path: str, token: str) -> dict | list:
    return _curl("GET", f"{DISCORD_API}{path}",
                 {"Authorization": f"Bot {token}"})


def discord_send(token: str, channel_id: str, text: str,
                 reply_to: str | None = None) -> bool:
    payload: dict = {"content": text}
    if reply_to:
        payload["message_reference"] = {
            "message_id": reply_to,
            "channel_id": channel_id,
            "fail_if_not_exists": False,
        }
    r = _curl("POST", f"{DISCORD_API}/channels/{channel_id}/messages",
              {"Authorization": f"Bot {token}",
               "Content-Type": "application/json"}, payload)
    if isinstance(r, dict) and r.get("id"):
        return True
    print(f"  ❌ Discord send: {r}")
    return False


# ═══════════════════════════════════════════════════════════════════
#  Feishu
# ═══════════════════════════════════════════════════════════════════

def fs_token(cfg: dict) -> str:
    global _fs_token, _fs_token_exp
    now = time.time()
    if _fs_token and now < _fs_token_exp - 120:
        return _fs_token
    fs = cfg["feishu"]
    r = _curl("POST", FEISHU_TOKEN_URL,
              {"Content-Type": "application/json"},
              {"app_id": fs["app_id"], "app_secret": fs["app_secret"]})
    if isinstance(r, dict) and r.get("code") != 0:
        raise RuntimeError(f"飞书 token: {r}")
    _fs_token = r["tenant_access_token"]
    _fs_token_exp = now + int(r.get("expire", 7200))
    return _fs_token


def fs_create(token: str, app_token: str, table_id: str, fields: dict):
    url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records"
    r = _curl("POST", url,
              {"Authorization": f"Bearer {token}",
               "Content-Type": "application/json"},
              {"fields": fields})
    if isinstance(r, dict) and r.get("code") != 0:
        raise RuntimeError(f"飞书写入: {r}")


def fs_batch_create(token: str, app_token: str, table_id: str,
                    records_fields: list[dict]):
    url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records/batch_create"
    batch = [{"fields": f} for f in records_fields]
    for i in range(0, len(batch), 100):
        chunk = batch[i:i + 100]
        r = _curl("POST", url,
                  {"Authorization": f"Bearer {token}",
                   "Content-Type": "application/json"},
                  {"records": chunk})
        if isinstance(r, dict) and r.get("code") != 0:
            raise RuntimeError(f"飞书批量写入: {r}")
        time.sleep(0.3)


# ═══════════════════════════════════════════════════════════════════
#  LLM (OpenRouter)
# ═══════════════════════════════════════════════════════════════════

def llm_generate(api_key: str, model: str,
                 system_prompt: str, user_prompt: str) -> str:
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.5,
    }
    r = _curl("POST", OPENROUTER_URL,
              {"Authorization": f"Bearer {api_key}",
               "Content-Type": "application/json"}, body, timeout=60)
    if isinstance(r, dict):
        return r.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
    return ""


# ═══════════════════════════════════════════════════════════════════
#  FETCH: pull historical messages from Discord API
# ═══════════════════════════════════════════════════════════════════

def _existing_message_ids() -> set[str]:
    return {str(r.get("msg_id", "")) for r in read_jsonl(DATA_DIR / "messages.jsonl")
            if r.get("msg_id")}


def _discord_api_team_mention(text: str, msg: dict, cfg: dict) -> bool:
    my_names = [u.lower().strip().lstrip("@")
                for u in (cfg.get("my_usernames") or []) if u]
    my_ids = [str(i).strip() for i in (cfg.get("my_discord_ids") or []) if i]
    text_low = text.lower()
    for u in my_names:
        if f"@{u}" in text_low or u in text_low:
            return True
    for uid in my_ids:
        if f"<@{uid}>" in text or f"<@!{uid}>" in text:
            return True
    for m in msg.get("mentions") or []:
        if str(m.get("id", "")) in my_ids:
            return True
    return False


def _discord_api_at_hazel(text: str, msg: dict, cfg: dict) -> bool:
    hazel_ids = [str(i).strip() for i in (cfg.get("hazel_discord_ids") or []) if i]
    hazel_names = [
        n.lower().strip().lstrip("@")
        for n in (cfg.get("hazel_usernames") or ["hazel"])
        if str(n).strip()
    ]
    for uid in hazel_ids:
        if f"<@{uid}>" in text or f"<@!{uid}>" in text:
            return True
    tl = text.lower()
    for n in hazel_names:
        if n and f"@{n}" in tl:
            return True
    for m in msg.get("mentions") or []:
        if str(m.get("id", "")) in hazel_ids:
            return True
        un = (m.get("username") or "").lower()
        for n in hazel_names:
            if n and un == n:
                return True
    return False


def _discord_api_match_keywords(text: str, kws: list) -> list[str]:
    if not text or not kws:
        return []
    matched = []
    text_low = text.lower()
    for kw in kws:
        kw = (kw or "").strip()
        if kw and kw.lower() in text_low:
            matched.append(kw)
    return matched


def do_fetch(
    cfg: dict,
    channel_ids: list[str],
    limit: int = 200,
    *,
    auto_ack: bool = False,
):
    """Fetch recent messages from specified Discord channels via REST API."""
    bot_token = cfg.get("discord_bot_token", "").strip()
    if not bot_token:
        print("❌ discord_bot_token 未配置"); return

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    existing = _existing_message_ids()
    total = 0
    skipped = 0
    skipped_bot = 0
    acked = 0
    ack_cfg = cfg.get("auto_ack") or {}
    ack_msg = (ack_cfg.get("message") or "").strip() or (
        "Thanks for your message — it has been logged and our team "
        "will look into it as soon as possible."
    )
    ack_mode = (ack_cfg.get("mode") or "all").strip().lower()

    for ch_id in channel_ids:
        ch_id = ch_id.strip()
        if not ch_id:
            continue

        print(f"\n📥 拉取频道 {ch_id} ...")
        ch_info = discord_get(f"/channels/{ch_id}", bot_token)
        ch_name = ch_info.get("name", ch_id) if isinstance(ch_info, dict) else ch_id
        ch_type = ch_info.get("type", 0) if isinstance(ch_info, dict) else 0
        guild_id = str(ch_info.get("guild_id", "") or "") if isinstance(ch_info, dict) else ""

        # For forum channels (type 15), fetch threads first
        if ch_type == 15:
            print(f"  论坛频道: #{ch_name}")
            threads = _fetch_forum_threads(ch_id, bot_token)
            for thread in threads:
                tid = str(thread["id"])
                t_name = thread.get("name", tid)
                msgs = _fetch_channel_messages(tid, bot_token, limit=limit)
                for m in msgs:
                    mid = str(m.get("id", ""))
                    if mid in existing:
                        skipped += 1
                        continue
                    rec = _discord_msg_to_record(
                        m, f"{ch_name}/{t_name}", tid, True, ch_id, guild_id, cfg,
                    )
                    if not rec.get("msg_id"):
                        skipped_bot += 1
                        continue
                    append_jsonl(DATA_DIR / "messages.jsonl", rec)
                    existing.add(mid)
                    total += 1
                    if auto_ack and _fetch_should_ack(rec, ack_mode):
                        body = ack_msg
                        if ack_cfg.get("mention_author", True) and rec.get("author_id"):
                            body = f"<@{rec['author_id']}> {body}"
                        if discord_send(bot_token, tid, body, reply_to=mid):
                            acked += 1
                        time.sleep(0.35)
                print(f"    帖子 [{t_name}]: {len(msgs)} 条消息")
                time.sleep(0.3)
        else:
            print(f"  文字频道: #{ch_name}")
            msgs = _fetch_channel_messages(ch_id, bot_token, limit=limit)
            for m in msgs:
                mid = str(m.get("id", ""))
                if mid in existing:
                    skipped += 1
                    continue
                rec = _discord_msg_to_record(
                    m, ch_name, ch_id, False, "", guild_id, cfg,
                )
                if not rec.get("msg_id"):
                    skipped_bot += 1
                    continue
                append_jsonl(DATA_DIR / "messages.jsonl", rec)
                existing.add(mid)
                total += 1
                if auto_ack and _fetch_should_ack(rec, ack_mode):
                    body = ack_msg
                    if ack_cfg.get("mention_author", True) and rec.get("author_id"):
                        body = f"<@{rec['author_id']}> {body}"
                    if discord_send(bot_token, ch_id, body, reply_to=mid):
                        acked += 1
                    time.sleep(0.35)
            print(f"  拉取 {len(msgs)} 条消息")

    print(f"\n✅ 新增写入 {total} 条（跳过已存在 {skipped}，跳过 bot 等 {skipped_bot}）"
          f"→ {DATA_DIR / 'messages.jsonl'}")
    if auto_ack:
        print(f"   auto-ack 已发送 {acked} 条回复")


def _fetch_should_ack(rec: dict, mode: str) -> bool:
    if mode == "all":
        return True
    if mode == "mention":
        return bool(rec.get("is_mention") or rec.get("at_hazel"))
    if mode == "hazel":
        return bool(rec.get("at_hazel"))
    return True


def _fetch_channel_messages(channel_id: str, token: str,
                            limit: int = 200) -> list[dict]:
    all_msgs = []
    before = None
    while len(all_msgs) < limit:
        batch = min(100, limit - len(all_msgs))
        path = f"/channels/{channel_id}/messages?limit={batch}"
        if before:
            path += f"&before={before}"
        msgs = discord_get(path, token)
        if not isinstance(msgs, list) or not msgs:
            break
        all_msgs.extend(msgs)
        if len(msgs) < batch:
            break
        before = msgs[-1]["id"]
        time.sleep(0.5)
    all_msgs.sort(key=lambda m: m["id"])
    return all_msgs


def _fetch_forum_threads(channel_id: str, token: str) -> list[dict]:
    all_threads = []
    ch_info = discord_get(f"/channels/{channel_id}", token)
    guild_id = ch_info.get("guild_id") if isinstance(ch_info, dict) else None
    if guild_id:
        active = discord_get(f"/guilds/{guild_id}/threads/active", token)
        if isinstance(active, dict):
            for t in active.get("threads") or []:
                if str(t.get("parent_id")) == str(channel_id):
                    all_threads.append(t)

    before = None
    while True:
        path = f"/channels/{channel_id}/threads/archived/public?limit=100"
        if before:
            path += f"&before={before}"
        data = discord_get(path, token)
        if not isinstance(data, dict):
            break
        threads = data.get("threads") or []
        if not threads:
            break
        all_threads.extend(threads)
        if not data.get("has_more"):
            break
        before = threads[-1].get("thread_metadata", {}).get("archive_timestamp")
        time.sleep(0.5)

    return all_threads


def _discord_msg_to_record(
    msg: dict,
    channel_name: str,
    channel_id: str,
    is_thread: bool,
    parent_forum_id: str,
    guild_id: str,
    cfg: dict,
) -> dict:
    author = msg.get("author", {})
    if author.get("bot"):
        return {}
    ts = msg.get("timestamp") or ""
    if ts:
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            ts = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
        except ValueError:
            pass
    text = msg.get("content", "")
    match_kws = cfg.get("match_keywords") or []
    matched = _discord_api_match_keywords(text, match_kws)
    return {
        "msg_id": str(msg.get("id", "")),
        "channel_id": channel_id,
        "channel_name": channel_name,
        "guild_id": guild_id,
        "is_thread": is_thread,
        "parent_id": parent_forum_id or "",
        "author_id": str(author.get("id", "")),
        "author_name": author.get("global_name") or author.get("username") or "",
        "author_username": author.get("username", ""),
        "text": text,
        "timestamp": ts,
        "is_mention": _discord_api_team_mention(text, msg, cfg),
        "at_hazel": _discord_api_at_hazel(text, msg, cfg),
        "matched_keywords": matched,
    }


# ═══════════════════════════════════════════════════════════════════
#  SCAN: scan stored messages for mentions / keywords → generate drafts
# ═══════════════════════════════════════════════════════════════════

def do_scan(cfg: dict):
    """Scan messages.jsonl, find matches, generate LLM drafts → mentions.jsonl."""
    my_names = [u.lower().strip().lstrip("@")
                for u in (cfg.get("my_usernames") or []) if u]
    my_ids = [str(i).strip() for i in (cfg.get("my_discord_ids") or []) if i]
    match_kws = cfg.get("match_keywords") or []
    llm_key = (cfg.get("openrouter_api_key") or "").strip()
    llm_model = cfg.get("openrouter_model", "openai/gpt-4o-mini")

    messages = read_jsonl(DATA_DIR / "messages.jsonl")
    existing_mentions = read_jsonl(DATA_DIR / "mentions.jsonl")
    seen_msg_ids = {m.get("msg_id") for m in existing_mentions if m.get("msg_id")}

    print(f"📊 共 {len(messages)} 条消息, {len(existing_mentions)} 条已有 mention")
    counter = 0
    new_mentions = []

    for rec in messages:
        mid = rec.get("msg_id", "")
        if mid in seen_msg_ids:
            continue

        text = rec.get("text", "")
        text_low = text.lower()

        is_mention = rec.get("is_mention")
        if is_mention is None:
            is_mention = any(f"@{u}" in text_low or u in text_low for u in my_names)
            for uid in my_ids:
                if f"<@{uid}>" in text or f"<@!{uid}>" in text:
                    is_mention = True

        ah = rec.get("at_hazel")
        if ah is None:
            at_hazel = _discord_api_at_hazel(text, {"mentions": []}, cfg)
        else:
            at_hazel = bool(ah)

        matched = [kw for kw in match_kws
                   if (kw or "").strip() and kw.lower() in text_low]

        if not is_mention and not matched and not at_hazel:
            continue

        counter += 1
        mention_id = f"dc_scan_{int(time.time())}_{counter}"

        triggers = []
        trigger_labels = []
        if is_mention:
            triggers.append("mention")
            trigger_labels.append("@提及")
        if at_hazel:
            triggers.append("hazel")
            trigger_labels.append("@Hazel/BD")
        if matched:
            triggers.append("keyword")
            trigger_labels.append(f"关键词({', '.join(matched)})")

        # LLM draft
        draft = ""
        if llm_key:
            print(f"  🤖 生成草稿: {text[:60]}...", end=" ", flush=True)
            sys_p = (
                "You are a DevRel engineer replying in a Discord channel. "
                "Write a concise, professional, friendly reply "
                "in the SAME LANGUAGE as the message."
            )
            kw_line = f"\nMatched keywords: {', '.join(matched)}" if matched else ""
            usr_p = (
                f"Channel: #{rec.get('channel_name', '?')}\n"
                f"Message:\n[{rec.get('author_name', '?')}]: {text}"
                f"{kw_line}\n\n"
                f"Generate a draft reply."
            )
            try:
                draft = llm_generate(llm_key, llm_model, sys_p, usr_p)
                print("✓")
            except Exception as e:
                print(f"✗ {e}")
                draft = f"[LLM error: {e}]"
            time.sleep(0.5)

        mention_rec = {
            "mention_id": mention_id,
            "msg_id": mid,
            "channel_id": rec.get("channel_id", ""),
            "channel_name": rec.get("channel_name", ""),
            "guild_id": rec.get("guild_id", ""),
            "is_thread": rec.get("is_thread", False),
            "parent_id": rec.get("parent_id", ""),
            "author_id": rec.get("author_id", ""),
            "author_name": rec.get("author_name", ""),
            "author_username": rec.get("author_username", ""),
            "text": text,
            "context": [],
            "draft_reply": draft,
            "status": "pending",
            "final_reply": "",
            "reply_timestamp": "",
            "timestamp": rec.get("timestamp", ""),
            "triggers": triggers,
            "trigger_label": " · ".join(trigger_labels),
            "matched_keywords": matched,
            "at_hazel": at_hazel,
        }
        new_mentions.append(mention_rec)
        append_jsonl(DATA_DIR / "mentions.jsonl", mention_rec)

    print(f"\n✅ 新增 {len(new_mentions)} 条待审核项")
    return new_mentions


# ═══════════════════════════════════════════════════════════════════
#  EXPORT: mentions.jsonl → review.xlsx
# ═══════════════════════════════════════════════════════════════════

HEADER_FILL = None
HEADER_FONT = None
APPROVE_FILL = None
SKIP_FILL = None
CELL_FONT = None
WRAP = None

COLS = [
    ("操作",          12),
    ("mention_id",    22),
    ("频道",          20),
    ("触发类型",      18),
    ("发送者",        16),
    ("用户名",        14),
    ("原始消息",      50),
    ("上下文",        40),
    ("AI草稿",        50),
    ("时间",          20),
    ("channel_id",    20),
    ("msg_id",        20),
]


def _init_styles():
    global HEADER_FILL, HEADER_FONT, APPROVE_FILL, SKIP_FILL, CELL_FONT, WRAP
    if HEADER_FILL is not None:
        return
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    APPROVE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    SKIP_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    CELL_FONT = Font(name="Arial", size=11)
    WRAP = Alignment(wrap_text=True, vertical="top")


def do_export():
    if openpyxl is None:
        print("❌ 需要 openpyxl:  pip install openpyxl"); return
    _init_styles()

    mentions = read_jsonl(DATA_DIR / "mentions.jsonl")
    pending = [m for m in mentions if m.get("status") == "pending"]

    if not pending:
        print("没有待审核项。")
        print("  提示: 先运行 python3 dc_review.py fetch --from-config 或 --channels <ID>")
        print(f"  再运行 python3 dc_review.py scan 扫描匹配项")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    for ci, (label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.freeze_panes = "A2"

    for ri, m in enumerate(pending, 2):
        ctx = m.get("context") or []
        ctx_str = "\n".join(ctx[-5:]) if ctx else ""
        row = [
            "",
            m.get("mention_id", ""),
            m.get("channel_name", ""),
            m.get("trigger_label", ""),
            m.get("author_name", ""),
            m.get("author_username", ""),
            m.get("text", ""),
            ctx_str,
            m.get("draft_reply", ""),
            m.get("timestamp", ""),
            m.get("channel_id", ""),
            m.get("msg_id", ""),
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP

        ws.row_dimensions[ri].height = max(
            60, 20 * max(1, len(str(row[6])) // 40))

    ws2 = wb.create_sheet("使用说明")
    for i, line in enumerate([
        "Discord DevRel Reviewer — 使用说明",
        "",
        "1. A列(操作) 填：approve / edit / skip / 留空",
        "2. 如果 edit，直接改 I列(AI草稿) 的内容",
        "3. 保存 → python3 dc_review.py send",
        "4. 发送后推飞书 → python3 dc_review.py sync",
        "",
        "⚠️ 不要改 K列(channel_id) 和 L列(msg_id)",
    ], 1):
        c = ws2.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=12 if i == 1 else 11,
                       bold=(i == 1))
    ws2.column_dimensions["A"].width = 55

    wb.save(EXCEL_PATH)
    print(f"✅ 导出 {len(pending)} 条 → {EXCEL_PATH}")
    print(f"   编辑 A列(操作) + I列(AI草稿) → 保存")
    print(f"   然后: python3 dc_review.py send")
    subprocess.run(["open", str(EXCEL_PATH)], check=False)


# ═══════════════════════════════════════════════════════════════════
#  SEND: review.xlsx → Discord batch reply
# ═══════════════════════════════════════════════════════════════════

def do_send(cfg: dict, dry_run: bool = False):
    if openpyxl is None:
        print("❌ 需要 openpyxl:  pip install openpyxl"); return
    _init_styles()

    bot_token = cfg.get("discord_bot_token", "").strip()
    if not bot_token and not dry_run:
        print("❌ discord_bot_token 未配置"); return
    if not EXCEL_PATH.is_file():
        print(f"❌ 找不到 {EXCEL_PATH}，先 export"); return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Review"]

    stats = {"approve": 0, "edit": 0, "skip": 0, "error": 0}

    for row in ws.iter_rows(min_row=2, values_only=False):
        if len(row) < len(COLS):
            continue
        action     = (row[0].value or "").strip().lower()
        mid        = (row[1].value or "").strip()
        ch_name    = row[2].value or ""
        draft      = row[8].value or ""
        orig_msg   = row[6].value or ""
        channel_id = str(row[10].value or "").strip()
        msg_id     = str(row[11].value or "").strip()

        if not action:
            continue

        if action == "skip":
            stats["skip"] += 1
            print(f"  ⏭ [#{ch_name}] {orig_msg[:50]}")
            row[0].fill = SKIP_FILL
            append_jsonl(REPLIES_LOG, {
                "mention_id": mid, "channel_name": ch_name,
                "text": orig_msg, "status": "skipped",
                "timestamp": datetime.now(timezone.utc).strftime(
                    "%Y-%m-%d %H:%M:%S UTC"),
            })
            continue

        if action in ("approve", "edit"):
            reply_text = draft
            reply_to_id = msg_id if msg_id else None

            if dry_run:
                print(f"  🔍 [{action}] → [#{ch_name}]: {reply_text[:80]}")
                stats[action] += 1
                continue

            ok = discord_send(bot_token, channel_id, reply_text,
                              reply_to=reply_to_id)
            if ok:
                stats[action] += 1
                row[0].fill = APPROVE_FILL
                print(f"  ✅ [#{ch_name}] → {reply_text[:60]}")
                append_jsonl(REPLIES_LOG, {
                    "mention_id": mid,
                    "channel_id": channel_id,
                    "channel_name": ch_name,
                    "text": orig_msg,
                    "status": "sent",
                    "final_reply": reply_text,
                    "reply_timestamp": datetime.now(
                        timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                })
            else:
                stats["error"] += 1
            time.sleep(0.5)

    if not dry_run:
        wb.save(EXCEL_PATH)

    tag = "🔍 DRY RUN" if dry_run else "📊"
    print(f"\n{tag} approve={stats['approve']} edit={stats['edit']} "
          f"skip={stats['skip']} error={stats['error']}")


# ═══════════════════════════════════════════════════════════════════
#  SYNC: local data → Feishu Bitable
# ═══════════════════════════════════════════════════════════════════

def do_sync(cfg: dict, tables: str = "all"):
    """
    Sync local JSONL to Feishu.
    tables: "all", "messages", "mentions", "replies"
    """
    fs = cfg.get("feishu") or {}
    if not fs.get("enabled"):
        print("⚠️  feishu.enabled=false"); return

    tk = fs_token(cfg)
    app_token = fs["app_token"]

    if tables in ("all", "messages"):
        msg_tid = fs.get("messages_table_id")
        if msg_tid:
            messages = read_jsonl(DATA_DIR / "messages.jsonl")
            print(f"📤 同步 {len(messages)} 条消息到飞书消息表 ...")
            fields_list = []
            for rec in messages:
                kw_str = ", ".join(rec.get("matched_keywords") or [])
                fields_list.append({
                    "频道": rec.get("channel_name", ""),
                    "发送者": rec.get("author_name", ""),
                    "用户名": rec.get("author_username", ""),
                    "消息内容": (rec.get("text") or "")[:2000],
                    "时间": rec.get("timestamp", ""),
                    "是否@我": "是" if rec.get("is_mention") else "否",
                    "匹配关键词": kw_str or "无",
                })
            try:
                fs_batch_create(tk, app_token, msg_tid, fields_list)
                print(f"  ✅ 消息: {len(fields_list)} 条")
            except Exception as e:
                print(f"  ⚠️ {e}")
        else:
            print("  ⚠️ messages_table_id 未配置，跳过消息表")

    if tables in ("all", "mentions"):
        review_tid = fs.get("review_table_id")
        if review_tid:
            mentions = read_jsonl(DATA_DIR / "mentions.jsonl")
            print(f"📤 同步 {len(mentions)} 条提及到飞书审核表 ...")
            ok = 0
            for m in mentions:
                ctx = m.get("context") or []
                try:
                    fs_create(tk, app_token, review_tid, {
                        "mention_id":   m.get("mention_id", ""),
                        "频道":         m.get("channel_name", ""),
                        "触发类型":     m.get("trigger_label", ""),
                        "发送者":       m.get("author_name", ""),
                        "用户名":       m.get("author_username", ""),
                        "原始消息":     (m.get("text") or "")[:2000],
                        "上下文":       "\n".join(ctx[-5:])[:2000],
                        "AI草稿":       (m.get("draft_reply") or "")[:2000],
                        "状态":         m.get("status", "pending"),
                        "channel_id":   m.get("channel_id", ""),
                        "msg_id":       m.get("msg_id", ""),
                        "时间":         m.get("timestamp", ""),
                    })
                    ok += 1
                except Exception as e:
                    print(f"  ⚠️ {e}")
            print(f"  ✅ 提及: {ok}/{len(mentions)}")
        else:
            print("  ⚠️ review_table_id 未配置，跳过审核表")

    if tables in ("all", "replies"):
        review_tid = fs.get("review_table_id")
        if review_tid and REPLIES_LOG.is_file():
            replies = read_jsonl(REPLIES_LOG)
            sent = [r for r in replies if r.get("status") == "sent"]
            if sent:
                print(f"📤 同步 {len(sent)} 条回复到飞书审核表 ...")
                ok = 0
                for r in sent:
                    try:
                        fs_create(tk, app_token, review_tid, {
                            "mention_id":   r.get("mention_id", ""),
                            "频道":         r.get("channel_name", ""),
                            "原始消息":     (r.get("text") or "")[:2000],
                            "最终回复":     (r.get("final_reply") or "")[:2000],
                            "状态":         "sent",
                            "操作":         "reply",
                            "时间":         r.get("reply_timestamp", ""),
                        })
                        ok += 1
                    except Exception as e:
                        print(f"  ⚠️ {e}")
                print(f"  ✅ 回复: {ok}/{len(sent)}")

    print("\n✅ 飞书同步完成")


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(
        description="Discord DevRel Reviewer (JSONL + Excel + Discord + Feishu)")
    sub = p.add_subparsers(dest="cmd")

    f = sub.add_parser("fetch", help="从 Discord API 拉取历史消息")
    f.add_argument("--channels", nargs="*", default=[],
                   help="频道/论坛 ID；可省略若使用 --from-config")
    f.add_argument("--from-config", action="store_true",
                   help="合并 config 中 watch_channels + watch_forums + fetch_extra_channel_ids")
    f.add_argument("--ack", action="store_true",
                   help="对符合条件的新消息发送 auto_ack 英文回复（见 auto_ack.mode）")
    f.add_argument("--limit", type=int, default=200,
                   help="每个频道最多拉取条数")

    sub.add_parser("scan", help="扫描消息 → 匹配 → 生成 LLM 草稿")

    sub.add_parser("export", help="mentions → review.xlsx")

    se = sub.add_parser("send", help="review.xlsx → Discord 批量回复")
    se.add_argument("--dry-run", action="store_true")

    sy = sub.add_parser("sync", help="本地数据 → 飞书多维表格")
    sy.add_argument("--tables", choices=["all", "messages", "mentions", "replies"],
                    default="all")

    args = p.parse_args()
    if not args.cmd:
        p.print_help(); return

    cfg = load_config()

    if args.cmd == "fetch":
        chs: list[str] = [str(x).strip() for x in (args.channels or []) if str(x).strip()]
        if args.from_config:
            extra = cfg.get("fetch_extra_channel_ids") or []
            merge = (
                list(cfg.get("watch_channels") or [])
                + list(cfg.get("watch_forums") or [])
                + list(extra)
            )
            chs = [str(x).strip() for x in merge if str(x).strip()] + chs
        chs = list(dict.fromkeys(chs))
        if not chs:
            print("❌ 请指定 --channels ID 或使用 --from-config"); return
        do_fetch(cfg, chs, args.limit, auto_ack=args.ack)
    elif args.cmd == "scan":
        do_scan(cfg)
    elif args.cmd == "export":
        do_export()
    elif args.cmd == "send":
        do_send(cfg, dry_run=args.dry_run)
    elif args.cmd == "sync":
        do_sync(cfg, tables=args.tables)


if __name__ == "__main__":
    main()
