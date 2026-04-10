#!/usr/bin/env python3
"""
TG DevRel Reviewer — Local JSON + Excel workflow

Flow:
  1. In TG bot private chat: /export → bot sends pending_mentions.json
  2. Download the file to this folder
  3. python3 tg_review.py export    → JSON → review.xlsx (auto-opens)
  4. Edit Excel: A列(操作) = approve/edit/skip, I列(AI草稿) = modify text
  5. python3 tg_review.py send      → Excel → TG reply
  6. python3 tg_review.py sync      → push to Feishu
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl:  pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
PENDING_JSON = SCRIPT_DIR / "pending_mentions.json"
DOWNLOADS_JSON = Path.home() / "Downloads" / "pending_mentions.json"
EXCEL_PATH = SCRIPT_DIR / "review.xlsx"
REPLIES_LOG = SCRIPT_DIR / "replies.jsonl"

TG_API = "https://api.telegram.org"
FEISHU_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
FEISHU_API = "https://open.feishu.cn/open-apis/bitable/v1"

_fs_token: str | None = None
_fs_token_exp: float = 0.0


# ═══════════════════════════════════════════════════════════════════
#  Config
# ═══════════════════════════════════════════════════════════════════

def load_config() -> dict:
    if CONFIG_PATH.is_file():
        return json.loads(CONFIG_PATH.read_text("utf-8"))
    alt = SCRIPT_DIR.parent / "astrbot_plugin_tg_assistant" / "config.json"
    if alt.is_file():
        return json.loads(alt.read_text("utf-8"))
    print(f"❌ 找不到 config.json"); sys.exit(1)


def append_jsonl(path: Path, rec: dict):
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ═══════════════════════════════════════════════════════════════════
#  HTTP helper
# ═══════════════════════════════════════════════════════════════════

def _curl(method: str, url: str, headers: dict,
          body: dict | None = None) -> dict:
    cmd = ["curl", "-s", "-X", method, url]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    if body is not None:
        cmd += ["-d", json.dumps(body)]
    p = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    if not p.stdout.strip():
        return {}
    return json.loads(p.stdout)


# ═══════════════════════════════════════════════════════════════════
#  TG send
# ═══════════════════════════════════════════════════════════════════

def tg_send(bot_token: str, chat_id: str, text: str,
            reply_to: int | None = None) -> bool:
    payload: dict = {"chat_id": chat_id, "text": text}
    if reply_to:
        payload["reply_parameters"] = {"message_id": reply_to}
    url = f"{TG_API}/bot{bot_token}/sendMessage"
    r = _curl("POST", url, {"Content-Type": "application/json"}, payload)
    if not r.get("ok"):
        print(f"  ❌ TG: {r.get('description', r)}")
        return False
    return True


# ═══════════════════════════════════════════════════════════════════
#  Feishu
# ═══════════════════════════════════════════════════════════════════

def fs_token(cfg: dict) -> str:
    global _fs_token, _fs_token_exp
    now = time.time()
    if _fs_token and now < _fs_token_exp - 120:
        return _fs_token
    fs = cfg["feishu"]
    r = _curl("POST", FEISHU_TOKEN_URL,
              {"Content-Type": "application/json"},
              {"app_id": fs["app_id"], "app_secret": fs["app_secret"]})
    if r.get("code") != 0:
        raise RuntimeError(f"飞书 token: {r}")
    _fs_token = r["tenant_access_token"]
    _fs_token_exp = now + int(r.get("expire", 7200))
    return _fs_token


def fs_create(token: str, app_token: str, table_id: str, fields: dict):
    url = f"{FEISHU_API}/apps/{app_token}/tables/{table_id}/records"
    r = _curl("POST", url,
              {"Authorization": f"Bearer {token}",
               "Content-Type": "application/json"},
              {"fields": fields})
    if r.get("code") != 0:
        raise RuntimeError(f"飞书写入: {r}")


# ═══════════════════════════════════════════════════════════════════
#  Export: pending_mentions.json → review.xlsx
# ═══════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
APPROVE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
CELL_FONT = Font(name="Arial", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")

COLS = [
    ("操作",          12),
    ("mention_id",    22),
    ("群组",          20),
    ("触发类型",      14),
    ("发送者",        16),
    ("用户名",        14),
    ("原始消息",      50),
    ("上下文",        40),
    ("AI草稿",        50),
    ("时间",          20),
    ("chat_id",       16),
    ("msg_id",        14),
]


def _find_pending_json() -> Path | None:
    if PENDING_JSON.is_file():
        return PENDING_JSON
    if DOWNLOADS_JSON.is_file():
        import shutil
        shutil.move(str(DOWNLOADS_JSON), str(PENDING_JSON))
        print(f"📂 自动从 ~/Downloads/ 移入 pending_mentions.json")
        return PENDING_JSON
    for f in sorted(Path.home().joinpath("Downloads").glob("pending_mentions*.json"),
                    key=lambda p: p.stat().st_mtime, reverse=True):
        import shutil
        shutil.move(str(f), str(PENDING_JSON))
        print(f"📂 自动从 ~/Downloads/{f.name} 移入")
        return PENDING_JSON
    return None


def do_export():
    src = _find_pending_json()
    if not src:
        print(f"❌ 找不到 pending_mentions.json")
        print(f"   请先在 Bot 私聊里发 /export，下载到 ~/Downloads/")
        return

    mentions = json.loads(src.read_text("utf-8"))
    if not mentions:
        print("文件里没有待审核项。"); return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    for ci, (label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.freeze_panes = "A2"

    for ri, m in enumerate(mentions, 2):
        ctx = m.get("context") or []
        ctx_str = "\n".join(ctx[-5:]) if ctx else ""
        row = [
            "",
            m.get("mention_id", ""),
            m.get("chat_title", ""),
            m.get("trigger_label", ""),
            m.get("sender_name", ""),
            f"@{m['sender_username']}" if m.get("sender_username") else "",
            m.get("text", ""),
            ctx_str,
            m.get("draft_reply", ""),
            m.get("timestamp", ""),
            m.get("chat_id", ""),
            m.get("msg_id", ""),
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP

        ws.row_dimensions[ri].height = max(
            60, 20 * max(1, len(str(row[6])) // 40))

    ws2 = wb.create_sheet("使用说明")
    for i, line in enumerate([
        "TG DevRel Reviewer — 使用说明",
        "",
        "1. A列(操作) 填：approve / edit / skip / 留空",
        "2. 如果 edit，直接改 I列(AI草稿) 的内容",
        "3. 保存 → python3 tg_review.py send",
        "4. 发送后推飞书 → python3 tg_review.py sync",
        "",
        "⚠️ 不要改 K列(chat_id) 和 L列(msg_id)",
    ], 1):
        c = ws2.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=12 if i == 1 else 11,
                       bold=(i == 1))
    ws2.column_dimensions["A"].width = 55

    wb.save(EXCEL_PATH)
    print(f"✅ 导出 {len(mentions)} 条 → {EXCEL_PATH}")
    print(f"   编辑 A列(操作) + I列(AI草稿) → 保存")
    print(f"   然后: python3 tg_review.py send")
    subprocess.run(["open", str(EXCEL_PATH)], check=False)


# ═══════════════════════════════════════════════════════════════════
#  Send: review.xlsx → TG reply
# ═══════════════════════════════════════════════════════════════════

def do_send(cfg: dict, dry_run: bool = False):
    bot_token = cfg.get("telegram_bot_token", "").strip()
    if not bot_token and not dry_run:
        print("❌ telegram_bot_token 未配置"); return
    if not EXCEL_PATH.is_file():
        print(f"❌ 找不到 {EXCEL_PATH}，先 export"); return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Review"]

    stats = {"approve": 0, "edit": 0, "skip": 0, "error": 0}

    for row in ws.iter_rows(min_row=2, values_only=False):
        if len(row) < len(COLS):
            continue
        action    = (row[0].value or "").strip().lower()
        mid       = (row[1].value or "").strip()
        title     = row[2].value or ""
        draft     = row[8].value or ""
        orig_msg  = row[6].value or ""
        chat_id   = str(row[10].value or "").strip()
        msg_id    = str(row[11].value or "").strip()

        if not action:
            continue

        if action == "skip":
            stats["skip"] += 1
            print(f"  ⏭ [{title}] {orig_msg[:50]}")
            row[0].fill = SKIP_FILL
            rec = {"mention_id": mid, "chat_title": title,
                   "text": orig_msg, "status": "skipped",
                   "timestamp": datetime.now(timezone.utc).strftime(
                       "%Y-%m-%d %H:%M:%S UTC")}
            append_jsonl(REPLIES_LOG, rec)
            continue

        if action in ("approve", "edit"):
            reply_text = draft
            reply_to = int(msg_id) if msg_id.isdigit() else None

            if dry_run:
                print(f"  🔍 [{action}] → [{title}]: {reply_text[:80]}")
                stats[action] += 1
                continue

            ok = tg_send(bot_token, chat_id, reply_text, reply_to=reply_to)
            if ok:
                stats[action] += 1
                row[0].fill = APPROVE_FILL
                print(f"  ✅ [{title}] → {reply_text[:60]}")
                rec = {
                    "mention_id": mid,
                    "chat_id": chat_id,
                    "chat_title": title,
                    "text": orig_msg,
                    "status": "sent",
                    "final_reply": reply_text,
                    "reply_timestamp": datetime.now(
                        timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                }
                append_jsonl(REPLIES_LOG, rec)
            else:
                stats["error"] += 1
            time.sleep(0.5)

    if not dry_run:
        wb.save(EXCEL_PATH)

    tag = "🔍 DRY RUN" if dry_run else "📊"
    print(f"\n{tag} approve={stats['approve']} edit={stats['edit']} "
          f"skip={stats['skip']} error={stats['error']}")


# ═══════════════════════════════════════════════════════════════════
#  Sync: local data → Feishu review table
# ═══════════════════════════════════════════════════════════════════

def _trigger_display(raw: str | None) -> str:
    if not raw:
        return "无"
    m = {"mention": "@提及", "bug": "Bug反馈"}
    return "+".join(m.get(p, p) for p in raw.split("+"))


def do_sync(cfg: dict):
    fs = cfg.get("feishu") or {}
    if not fs.get("enabled"):
        print("⚠️  feishu.enabled=false"); return

    tk = fs_token(cfg)
    app_token = fs["app_token"]
    review_tid = fs.get("review_table_id")

    if not review_tid:
        print("⚠️  review_table_id 未配置"); return

    # Sync pending mentions
    if PENDING_JSON.is_file():
        mentions = json.loads(PENDING_JSON.read_text("utf-8"))
        print(f"📤 同步 {len(mentions)} 条提及到飞书审核表 ...")
        ok = 0
        for m in mentions:
            ctx = m.get("context") or []
            try:
                fs_create(tk, app_token, review_tid, {
                    "mention_id":   m.get("mention_id", ""),
                    "群组":         m.get("chat_title", ""),
                    "触发类型":     m.get("trigger_label", ""),
                    "发送者":       m.get("sender_name", ""),
                    "用户名":       f"@{m['sender_username']}" if m.get("sender_username") else "",
                    "原始消息":     (m.get("text") or "")[:2000],
                    "上下文":       "\n".join(ctx[-5:])[:2000],
                    "AI草稿":       (m.get("draft_reply") or "")[:2000],
                    "状态":         m.get("status", "pending"),
                    "chat_id":      m.get("chat_id", ""),
                    "msg_id":       m.get("msg_id", ""),
                    "时间":         m.get("timestamp", ""),
                })
                ok += 1
            except Exception as e:
                print(f"  ⚠️ {e}")
        print(f"  ✅ 提及: {ok}/{len(mentions)}")

    # Sync replies
    if REPLIES_LOG.is_file():
        replies = []
        with REPLIES_LOG.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        replies.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
        sent = [r for r in replies if r.get("status") == "sent"]
        if sent:
            print(f"📤 同步 {len(sent)} 条回复到飞书审核表 ...")
            ok = 0
            for r in sent:
                try:
                    fs_create(tk, app_token, review_tid, {
                        "mention_id":   r.get("mention_id", ""),
                        "群组":         r.get("chat_title", ""),
                        "原始消息":     (r.get("text") or "")[:2000],
                        "最终回复":     (r.get("final_reply") or "")[:2000],
                        "状态":         "sent",
                        "操作":         "reply",
                        "时间":         r.get("reply_timestamp", ""),
                    })
                    ok += 1
                except Exception as e:
                    print(f"  ⚠️ {e}")
            print(f"  ✅ 回复: {ok}/{len(sent)}")

    print("\n✅ 飞书同步完成")


# ═══════════════════════════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description="TG DevRel Reviewer (Local JSON + Excel)")
    sub = p.add_subparsers(dest="cmd")

    sub.add_parser("export", help="pending_mentions.json → review.xlsx")

    se = sub.add_parser("send", help="review.xlsx → TG 回复")
    se.add_argument("--dry-run", action="store_true")

    sub.add_parser("sync", help="本地数据 → 飞书审核表")

    args = p.parse_args()
    if not args.cmd:
        p.print_help(); return

    cfg = load_config()

    if args.cmd == "export":
        do_export()
    elif args.cmd == "send":
        do_send(cfg, dry_run=args.dry_run)
    elif args.cmd == "sync":
        do_sync(cfg)


if __name__ == "__main__":
    main()
